# -*- coding: utf-8 -*-
"""Реестр описаний: что видит игрок, что скрыто, что кривое. Тест-киты."""
from __future__ import annotations

import re
import sys
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parents[1]
OUT = ROOT / "описания_способностей.xlsx"
sys.path.insert(0, str(Path(__file__).resolve().parent))
from build_full_review import load_all  # noqa: E402

INK = "1A1A1A"
HEAD = "2C2416"
HEAD_F = "F7F1E3"
YEL = "FFF3B0"
RED = "F4D6D0"
OK = "DFF0D8"
HID = "EDE4F5"
THIN = Border(
    left=Side(style="thin", color="D0C4A8"),
    right=Side(style="thin", color="D0C4A8"),
    top=Side(style="thin", color="D0C4A8"),
    bottom=Side(style="thin", color="D0C4A8"),
)
WRAP = Alignment(wrap_text=True, vertical="top")
FONT = Font(name="Calibri", size=10, color=INK)
FONT_B = Font(name="Calibri", size=10, bold=True, color=INK)
FONT_H = Font(name="Calibri", size=10, bold=True, color=HEAD_F)
FONT_T = Font(name="Calibri", size=16, bold=True, color=HEAD)


def fill(h):
    return PatternFill("solid", fgColor=h)


def n(v):
    if v is None:
        return None
    try:
        x = float(v)
        return int(x) if x == int(x) else x
    except (TypeError, ValueError):
        return v


def flat_of(ab):
    if ab.get("fl") is not None:
        return n(ab["fl"])
    if ab.get("flat") is not None:
        return n(ab["flat"])
    return None


def desc_looks_bare(raw):
    s = (raw or "").strip()
    if not s:
        return True
    leftover = re.sub(
        r"\b(кд|ход|хода|ходу|раунд|раунда|раундов|ярость|ярости|мана|маны|пар|пара|"
        r"энергия|энергии|очко|очка|очков|заряд|заряда|зарядов|серия|серии|область|"
        r"перезарядка|бесплатно|aoe|ст)\b",
        " ",
        s,
        flags=re.I,
    )
    leftover = re.sub(r"[\d\s·.,:;+\-−~=%/|+×xх()]+", " ", leftover)
    leftover = re.sub(r"\bт\b", " ", leftover, flags=re.I).strip()
    return len(leftover) < 4


def facts(ab):
    bits = []
    fl = flat_of(ab)
    t = ab.get("t") or ""
    if fl is not None:
        bits.append(f"кнопка {fl}т ({t or '?'})")
    if ab.get("p") not in (None, 1, 1.0):
        bits.append(f"power {ab.get('p')}")
    if ab.get("hits") and n(ab.get("hits")) > 1:
        bits.append(f"ударов {n(ab.get('hits'))}")
    ad = ab.get("applyDot") if isinstance(ab.get("applyDot"), dict) else None
    if ad:
        bits.append(f"дот «{ad.get('name') or ab.get('id')}» {n(ad.get('flat'))}т×{n(ad.get('turns'))}")
    ah = ab.get("applyHot") if isinstance(ab.get("applyHot"), dict) else None
    if ah:
        if ah.get("hpPct") is not None:
            bits.append(f"хот «{ah.get('name') or ab.get('id')}» {round(float(ah['hpPct'])*100)}% HP×{n(ah.get('turns'))}")
        else:
            bits.append(f"хот «{ah.get('name') or ab.get('id')}» {n(ah.get('flat'))}т×{n(ah.get('turns'))}")
    if ab.get("atkMod"):
        bits.append(f"+{round(float(ab['atkMod'])*100)}% атаки {n(ab.get('bt') or ab.get('buffTurns'))}х")
    if ab.get("petAtkMod"):
        bits.append(f"+{round(float(ab['petAtkMod'])*100)}% пета {n(ab.get('bt') or ab.get('buffTurns'))}х")
    if ab.get("dmgReduce") or ab.get("dr"):
        bits.append(f"−{round(float(ab.get('dmgReduce') or ab.get('dr') or 0)*100)}% входа {n(ab.get('bt') or ab.get('buffTurns'))}х")
    if ab.get("lifesteal"):
        bits.append(f"вампиризм {round(float(ab['lifesteal'])*100)}%")
    if ab.get("healFromDealt"):
        bits.append(f"хил {round(float(ab['healFromDealt'])*100)}% от нанесённого")
    if ab.get("vuln") and isinstance(ab.get("vuln"), dict):
        bits.append(f"уязв +{round(float(ab['vuln'].get('amount') or 0)*100)}% {n(ab['vuln'].get('turns'))}х")
    gb = ab.get("grantSelfBuff") if isinstance(ab.get("grantSelfBuff"), dict) else None
    if gb and gb.get("name"):
        bits.append(f"бафф «{gb.get('name')}»")
    if ab.get("fa") or ab.get("freeAction"):
        bits.append("без хода")
    if ab.get("ch"):
        bits.append(f"{n(ab.get('ch'))} заряда")
    if ab.get("cd"):
        bits.append(f"КД {n(ab.get('cd'))}")
    return bits


def propose(ab):
    bits = facts(ab)
    t = ab.get("t") or ""
    head = {
        "damage": "Удар по выбранной цели.",
        "aoe": "Урон по всем врагам.",
        "dot": "Накладывает периодический урон.",
        "heal": "Исцеляет выбранного союзника.",
        "heal_aoe": "Исцеляет отряд.",
        "shield": "Щит, поглощает урон.",
        "buff": "Усиливает вас.",
        "debuff": "Ослабляет врага.",
        "summon": "Призывает питомца.",
        "interrupt": "Сбивает каст, немота.",
        "taunt": "Перетягивает внимание.",
        "cc": "Контроль цели.",
        "cleanse": "Снимает пошатывание.",
    }.get(t, "")
    raw = (ab.get("d") or ab.get("desc") or "").strip()
    body = raw if raw and not desc_looks_bare(raw) else head
    extra = [f for f in bits if not f.startswith("кнопка") and not f.startswith("КД")]
    return " ".join(x for x in [body] + extra if x).strip()


def flags_of(ab, desc_now):
    flags = []
    raw = (ab.get("d") or ab.get("desc") or "").strip()
    if not raw:
        flags.append("пустое desc")
    elif desc_looks_bare(raw):
        flags.append("desc = голые цифры")
    has_child = bool(ab.get("applyDot") or ab.get("applyHot") or ab.get("vuln") or ab.get("lifesteal") or ab.get("healFromDealt"))
    text = (raw + " " + desc_now).lower()
    if has_child and not re.search(r"\d+\s*т|\d+\s*%", text):
        flags.append("эффект без цифр в тексте")
    ad = ab.get("applyDot") if isinstance(ab.get("applyDot"), dict) else None
    if ad and ad.get("name"):
        blob = raw + desc_now
        if f"«{ad.get('name')}»" not in blob and (ad.get("name") or "") not in blob:
            flags.append("дот не назван")
    if (ab.get("fa") or ab.get("freeAction")) and not re.search(r"без хода|не тратит|не заверш", text):
        flags.append("без хода не сказано")
    if not flags:
        flags.append("ок")
    return flags


HIDDEN = [
    ["Монах", "Хмелевар", "Дар хмелевара", "хот", "Когда хмелевара лечат: шанс = крит лечащего → 75% хила на 5 раундов.", "не было кнопки и пассивки", "пассивка + словарь"],
    ["Монах", "Хмелевар", "Пошатывание", "пул", "≈35% входящего в пул, тик по себе. Очищающий снимает долю в щит неуловимости.", "нет своей кнопки", "словарь + текст очищающего"],
    ["Монах", "Ткач туманов", "Носители тумана", "хил", "70% урона ткача → хил носителям Заживляющего.", "не было в пассивках", "пассивка"],
    ["Монах", "Ткач туманов", "Тик змеи", "пет", "3т хил + 3т урон после хода каждого героя и моба.", "не на кнопке подробно", "пассивка + текст змеи"],
    ["Жрец", "Послушание", "Искупление", "хил от урона", "Щит / молитва / Кара / огонь / исповедь во врага / пет кормят носителей.", "не было в пассивках", "пассивка + словарь"],
    ["Паладин", "Свет", "Выбор света", "хот", "Хил раненого → хот 2 р. от искусности.", "есть пассивка", "ок"],
    ["Друид", "Баланс", "Затмение", "бафф", "+20% атаки 3 хода при полной шкале.", "не было в пассивках", "пассивка"],
    ["Воин", "Оружие", "Кровотечение", "дот", "5т×4 с Смертельного / Колосса / Героического.", "на кнопке было пусто", "пассивка + «Кровотечение» в подсказке"],
    ["Воин", "Неистовство", "Необузданная ярость", "стаки", "Спендер ярости копит +% урона, без траты сбрасывает.", "есть пассивка", "ок"],
    ["Инженер", "Механист", "Делит удар", "перенос", "50% урона хозяина на бота.", "есть пассивка", "ок"],
    ["Разбойник", "все", "Кривая серии", "множитель", "0.22 / 0.42 / 0.68 / 1.05 / 1.55. На кнопке «при 5 очках».", "не было пассивки", "пассивка + строка в финишере"],
    ["Друид", "Сила зверя", "Кривая серии", "множитель", "Та же таблица, что у разбойника. Дикий рёв серию не ест.", "не было пассивки", "пассивка"],
    ["Рыцарь смерти", "Кровь", "Удар смерти от полученного", "хил", "Плюс доля полученного за последние 2 хода.", "на кнопке часто голо", "проверить desc"],
    ["Маг", "Огонь", "Раскалённая глыба", "окно", "Крит Огненного шара: следующая глыба 10 маны и 90т.", "скрыто в движке", "пассивка + словарь"],
    ["Маг", "Огонь", "Раскалённый столб", "окно", "Огненный столб 33%: следующий шар критует.", "скрыто в движке", "пассивка + словарь"],
    ["Маг", "Тайная магия", "Стаки вспышки", "ресурс", "Повтор вспышки +4 маны, обстрел ест стаки.", "частично в desc", "в desc вспышки"],
    ["Маг", "Лёд", "Копьё — область", "прок", "20% со стрелы: следующее копьё по области.", "скрыто в движке", "пассивка + словарь"],
    ["Шаман", "Стихии", "Выброс по шоку", "крит", "Гарантированный крит, если на цели Огненный шок.", "в desc коротко", "словарь"],
    ["Чернокнижник", "Колдовство", "Дух / Хватка за дот", "множитель", "+15% / +10% за свой дот на цели.", "в desc кнопок", "ок если desc жив"],
]


def parse_passives(text):
    rows = []
    for m in re.finditer(r"list\.push\(\{([\s\S]*?)\}\)", text):
        block = m.group(1)
        head = text[max(0, m.start() - 400):m.start()]
        cid = "?"
        sid = "?"
        cm = list(re.finditer(r"classId === '([a-z]+)'", head))
        sm = list(re.finditer(r"specId === '([a-z_]+)'", head))
        if cm:
            cid = cm[-1].group(1)
        if "classId === 'rogue'" in head and "specId ===" not in head[-160:]:
            sid = "все"
        elif "classId === 'paladin'" in head and "specId ===" not in head[-160:]:
            sid = "все"
        elif sm:
            sid = sm[-1].group(1)
        name = re.search(r"name:\s*'([^']+)'", block)
        detail = re.search(r"detail:\s*'([^']*)'", block)
        short = re.search(r"short:\s*'([^']*)'", block)
        pid = re.search(r"id:\s*'([^']+)'", block)
        rows.append({
            "class": cid,
            "spec": sid,
            "id": pid.group(1) if pid else "",
            "name": name.group(1) if name else "",
            "short": short.group(1) if short else "",
            "detail": detail.group(1) if detail else "",
        })
    return rows


def parse_glossary(text):
    rows = []
    m = re.search(r"const EFFECT_GLOSSARY = \{([\s\S]*?)\n  \};", text)
    if not m:
        return rows
    for km in re.finditer(r"'([^']+)':\s*'([^']*)'", m.group(1)):
        rows.append((km.group(1), km.group(2)))
    return rows


def style_head(ws, cols):
    for i in range(1, cols + 1):
        c = ws.cell(1, i)
        c.fill = fill(HEAD)
        c.font = FONT_H
        c.alignment = Alignment(wrap_text=True, vertical="center")
        c.border = THIN
    ws.row_dimensions[1].height = 28
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(cols)}1"


def paint_row(ws, ri, cols, color):
    for ci in range(1, cols + 1):
        cell = ws.cell(ri, ci)
        cell.font = FONT
        cell.alignment = WRAP
        cell.border = THIN
        if color:
            cell.fill = fill(color)


def main():
    kits = load_all()
    pas_src = (ROOT / "js" / "systems" / "passives.js").read_text(encoding="utf-8")
    ad_src = (ROOT / "js" / "systems" / "ability-data.js").read_text(encoding="utf-8")
    passives = parse_passives(pas_src)
    glossary = parse_glossary(ad_src)

    wb = Workbook()
    stats = {"abs": 0, "empty": 0, "bare": 0, "ok": 0, "fx": 0}

    ws = wb.active
    ws.title = "Как читать"

    ws_ab = wb.create_sheet("Способности")
    ws_ab.append([
        "Класс", "Спек", "id", "Имя", "Тип",
        "desc в файле", "Факты из данных", "Что видит игрок (слой подсказки)", "Флаги",
    ])
    style_head(ws_ab, 9)
    ri = 2
    effect_rows = []
    for c in kits:
        if c["meta"]["id"] == "cheat":
            continue
        for s in c["specs"]:
            for ab in s["abilities"]:
                fact = facts(ab)
                prop = propose(ab)
                flg = flags_of(ab, prop)
                raw = (ab.get("d") or ab.get("desc") or "").strip()
                ws_ab.append([
                    c["meta"]["name"], s["name"], ab.get("id") or "",
                    ab.get("n") or "", ab.get("t") or "",
                    raw,
                    " · ".join(str(x) for x in fact),
                    prop,
                    "; ".join(flg),
                ])
                color = None
                if "пустое desc" in flg or "эффект без цифр в тексте" in flg:
                    color = RED
                elif "desc = голые цифры" in flg or "дот не назван" in flg or "без хода не сказано" in flg:
                    color = YEL
                elif flg == ["ок"]:
                    color = OK
                paint_row(ws_ab, ri, 9, color)
                stats["abs"] += 1
                if "пустое desc" in flg:
                    stats["empty"] += 1
                if "desc = голые цифры" in flg:
                    stats["bare"] += 1
                if flg == ["ок"]:
                    stats["ok"] += 1
                ad = ab.get("applyDot") if isinstance(ab.get("applyDot"), dict) else None
                if ad:
                    effect_rows.append([
                        c["meta"]["name"], s["name"], ab.get("n"), "дот",
                        ad.get("name") or ab.get("id"),
                        f"{n(ad.get('flat'))}т за раунд, {n(ad.get('turns'))} р.",
                        "Наведи на «" + str(ad.get("name") or "") + "» в подсказке способности или на иконку у портрета.",
                    ])
                ah = ab.get("applyHot") if isinstance(ab.get("applyHot"), dict) else None
                if ah:
                    if ah.get("hpPct") is not None:
                        txt = f"{round(float(ah['hpPct'])*100)}% HP × {n(ah.get('turns'))} р."
                    else:
                        txt = f"{n(ah.get('flat'))}т × {n(ah.get('turns'))} р."
                    effect_rows.append([
                        c["meta"]["name"], s["name"], ab.get("n"), "хот",
                        ah.get("name") or ab.get("id"), txt,
                        "Наведи на имя хота в подсказке или на иконку у портрета.",
                    ])
                gb = ab.get("grantSelfBuff") if isinstance(ab.get("grantSelfBuff"), dict) else None
                if gb and gb.get("name"):
                    effect_rows.append([
                        c["meta"]["name"], s["name"], ab.get("n"), "бафф",
                        gb.get("name"), gb.get("tip") or "стак / окно",
                        "Наведи на «" + str(gb.get("name")) + "» в подсказке.",
                    ])
                ri += 1
    stats["fx"] = len(effect_rows)
    for i, w in enumerate([14, 18, 16, 28, 10, 42, 44, 56, 30], 1):
        ws_ab.column_dimensions[get_column_letter(i)].width = w
    ws_ab.auto_filter.ref = f"A1:I{ri-1}"

    ws["A1"] = "Описания способностей — реестр"
    ws["A1"].font = FONT_T
    ws.merge_cells("A1:B1")
    intro = [
        ("Лист «Способности»", "Каждая кнопка: сырой desc из файла, факты из данных, какой текст собирает слой подсказки, флаги."),
        ("Лист «Эффекты»", "Именные доты / хоты / баффы с кнопок. На это наводишь в кавычках «…» или на иконку у портрета."),
        ("Лист «Словарь»", "Что раскрывается при наведении на жёлтую «строку» в подсказке."),
        ("Лист «Скрытое»", "Проки движка без своей кнопки (Дар хмелевара, тик змеи, крит шара…)."),
        ("Лист «Пассивки»", "Что лежит в кармане пассивок (разбор passives.js)."),
        ("Цвет", "Жёлтый — голый/кривой desc в файле. Красный — пусто или эффект без цифр. Фиолетовый — скрытое. Зелёный — ок."),
        ("В игре", "Наведи на способность. Жёлтые «Кавычки» можно навести ещё раз — снизу раскроется, как эффект работает. То же на иконке у портрета."),
        ("Слой, не файл", "Пустой desc в ките не значит «игрок ничего не видит»: подсказка собирает фразу из типа + цифр дота/хота. Файл специально часто пустой."),
        ("Не трогал", "Цифры баланса и основу. Только тексты, словарь, пассивки и этот файл."),
        ("Кнопок в реестре", str(stats["abs"])),
        ("Пустой desc в файле", str(stats["empty"])),
        ("Desc = голые цифры", str(stats["bare"])),
        ("Без флагов (ок)", str(stats["ok"])),
        ("Именных эффектов", str(stats["fx"])),
        ("Статей словаря", str(len(glossary))),
        ("Пассивок в коде", str(len(passives))),
    ]
    ws["A3"] = "Что"
    ws["B3"] = "Зачем"
    style_head(ws, 2)
    ws["A3"].fill = fill(HEAD)
    for i, (a, b) in enumerate(intro, 4):
        ws.cell(i, 1, a).font = FONT_B
        ws.cell(i, 2, b).font = FONT
        ws.cell(i, 1).alignment = WRAP
        ws.cell(i, 2).alignment = WRAP
        ws.cell(i, 1).border = THIN
        ws.cell(i, 2).border = THIN
        ws.row_dimensions[i].height = 30
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 110

    ws = wb.create_sheet("Эффекты")
    ws.append(["Класс", "Спек", "Способность", "Вид", "Имя эффекта", "Цифры", "Как открыть"])
    style_head(ws, 7)
    for i, row in enumerate(effect_rows, 2):
        ws.append(row)
        paint_row(ws, i, 7, HID)
    for i, w in enumerate([14, 18, 26, 8, 28, 30, 52], 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    if effect_rows:
        ws.auto_filter.ref = f"A1:G{len(effect_rows)+1}"

    ws = wb.create_sheet("Словарь")
    ws.append(["Имя", "Текст при наведении", "Есть в эффектах кнопок"])
    style_head(ws, 3)
    fx_names = {str(r[4]) for r in effect_rows}
    for i, (name, text) in enumerate(glossary, 2):
        ws.append([name, text, "да" if name in fx_names else "только словарь / скрытое"])
        paint_row(ws, i, 3, OK if name in fx_names else HID)
        ws.row_dimensions[i].height = 36
    for i, w in enumerate([28, 90, 28], 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws = wb.create_sheet("Скрытое")
    ws.append(["Класс", "Спек", "Имя", "Вид", "Что делает", "Почему скрыто", "Статус этой ночью"])
    style_head(ws, 7)
    for i, row in enumerate(HIDDEN, 2):
        ws.append(row)
        last = str(row[-1])
        color = OK if ("пассивка" in last or "словарь" in last or last.startswith("ок")) else RED
        paint_row(ws, i, 7, color)
        ws.row_dimensions[i].height = 40
    for i, w in enumerate([18, 16, 28, 14, 62, 30, 32], 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws = wb.create_sheet("Пассивки")
    ws.append(["Класс (id)", "Спек (id)", "id", "Имя", "Коротко", "Полный текст"])
    style_head(ws, 6)
    for i, p in enumerate(passives, 2):
        ws.append([p["class"], p["spec"], p["id"], p["name"], p["short"], p["detail"]])
        paint_row(ws, i, 6, OK)
        ws.row_dimensions[i].height = 36
    for i, w in enumerate([16, 16, 22, 28, 22, 70], 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    if passives:
        ws.auto_filter.ref = f"A1:F{len(passives)+1}"

    wb.save(OUT)
    print("wrote", OUT)
    print("abilities", stats["abs"], "empty", stats["empty"], "bare", stats["bare"], "ok", stats["ok"])
    print("effects", stats["fx"], "glossary", len(glossary), "passives", len(passives))


if __name__ == "__main__":
    main()
