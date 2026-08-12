# -*- coding: utf-8 -*-
"""Полный обзор баланса Тест: факты из class-balance + предложения + колонка задач."""
from __future__ import annotations

import re
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parents[1]
OUT = ROOT / "баланс_полный_обзор.xlsx"
BAL = ROOT / "class-balance"
STATE = ROOT / "js" / "state.js"

FLAT_REF = 15
SCALE8 = 1 + (8 - 2) * 0.015
SCALE15 = 1 + (15 - 2) * 0.015
AVG_SC = (SCALE8 + SCALE15) / 2

CLASS_RU = {
    "warrior": "Воин", "paladin": "Паладин", "hunter": "Охотник", "rogue": "Разбойник",
    "priest": "Жрец", "deathknight": "Рыцарь смерти", "shaman": "Шаман", "mage": "Маг",
    "warlock": "Чернокнижник", "monk": "Монах", "druid": "Друид", "engineer": "Гном-инженер",
}
ROLE_RU = {"tank": "танк", "healer": "хил", "dps": "дд"}
OWNER_YOU = {
    "warrior:arms", "warrior:fury", "warrior:protection",
    "paladin:holy", "paladin:protection", "paladin:retribution",
    "shaman:restoration", "warlock:demonology", "monk:brewmaster", "engineer:tinkerer",
}
DOT_FORCE = {
    "moonfire", "sunfire", "rake", "rip", "lacerate", "rend", "agony", "ua", "corruption",
    "immolate", "doom", "swp", "vt", "devouring", "holy_fire", "serpent", "black_arrow",
    "garrote", "rupture", "living_bomb", "nether_tempest", "flame_shock", "plague_strike",
    "outbreak", "poison", "d", "dot", "sticky_bomb",
}

# ── styles ──
INK = "1A1A1A"
GOLD = "C4A35A"
BG = "F7F1E3"
HEAD = "2C2416"
HEAD_F = "F7F1E3"
YEL = "FFF3B0"
FACT = "E8F0E3"
PROP = "E3EEF7"
P0 = "F4D6D0"
P1 = "F8E6C8"
ENG = "EDE4F5"
THIN = Border(
    left=Side(style="thin", color="D0C4A8"),
    right=Side(style="thin", color="D0C4A8"),
    top=Side(style="thin", color="D0C4A8"),
    bottom=Side(style="thin", color="D0C4A8"),
)
WRAP = Alignment(wrap_text=True, vertical="top")
FONT = Font(name="Calibri", size=10, color=INK)
FONT_B = Font(name="Calibri", size=10, bold=True, color=INK)
FONT_H = Font(name="Calibri", size=10, bold=True, color=HEAD_F)
FONT_T = Font(name="Calibri", size=16, bold=True, color=HEAD)
FONT_S = Font(name="Calibri", size=11, italic=True, color="4A4030")


def fill(hex_):
    return PatternFill("solid", fgColor=hex_)


def parse_js_val(s, i):
    while i < len(s) and s[i] in " \n\r\t":
        i += 1
    if i >= len(s):
        return None, i
    c = s[i]
    if c in ("'", '"'):
        q = c
        i += 1
        buf = []
        while i < len(s):
            if s[i] == "\\":
                buf.append(s[i + 1] if i + 1 < len(s) else "")
                i += 2
                continue
            if s[i] == q:
                return "".join(buf), i + 1
            buf.append(s[i])
            i += 1
        return "".join(buf), i
    if c == "{":
        depth = 0
        start = i
        while i < len(s):
            if s[i] == "{":
                depth += 1
            elif s[i] == "}":
                depth -= 1
                if depth == 0:
                    return parse_js_obj(s[start : i + 1]), i + 1
            elif s[i] in ("'", '"'):
                qq = s[i]
                i += 1
                while i < len(s) and s[i] != qq:
                    if s[i] == "\\":
                        i += 1
                    i += 1
            i += 1
        return {}, i
    if s.startswith("true", i) and (i + 4 == len(s) or not s[i + 4].isalnum()):
        return True, i + 4
    if s.startswith("false", i) and (i + 5 == len(s) or not s[i + 5].isalnum()):
        return False, i + 5
    if s.startswith("null", i):
        return None, i + 4
    m = re.match(r"-?\d+\.?\d*", s[i:])
    if m:
        t = m.group(0)
        return (float(t) if "." in t else int(t)), i + len(t)
    m = re.match(r"[A-Za-z_][\w.]*", s[i:])
    if m:
        return m.group(0), i + len(m.group(0))
    return None, i + 1


def parse_js_obj(blob):
    if not blob or blob[0] != "{":
        return {}
    inner = blob[1:-1]
    out = {}
    i = 0
    n = len(inner)
    while i < n:
        while i < n and inner[i] in " \n\r\t,":
            i += 1
        if i >= n:
            break
        if inner[i] in ("'", '"'):
            key, i = parse_js_val(inner, i)
        else:
            m = re.match(r"[\w$]+", inner[i:])
            if not m:
                i += 1
                continue
            key = m.group(0)
            i += len(key)
        while i < n and inner[i] in " \n\r\t":
            i += 1
        if i >= n or inner[i] != ":":
            continue
        i += 1
        val, i = parse_js_val(inner, i)
        out[key] = val
    return out


def extract_A_blocks(text):
    out = []
    for m in re.finditer(r"\bA\s*\(\s*\{", text):
        start = text.find("{", m.start())
        depth = 0
        j = start
        while j < len(text):
            if text[j] == "{":
                depth += 1
            elif text[j] == "}":
                depth -= 1
                if depth == 0:
                    out.append(parse_js_obj(text[start : j + 1]))
                    break
            elif text[j] in ("'", '"'):
                q = text[j]
                j += 1
                while j < len(text) and text[j] != q:
                    if text[j] == "\\":
                        j += 1
                    j += 1
            j += 1
    return out


def split_specs(text):
    """Cut file into spec chunks by id + name + role nearby."""
    chunks = []
    pat = re.compile(
        r"id:\s*'([a-z_]+)'\s*,\s*(?:\n\s*)?name:\s*'([^']+)'\s*,\s*(?:\n\s*)?nameEn:\s*'([^']+)'\s*,\s*(?:\n\s*)?role:\s*'([a-z]+)'",
        re.M,
    )
    for m in pat.finditer(text):
        chunks.append((m.start(), m.group(1), m.group(2), m.group(3), m.group(4)))
    specs = []
    for i, (pos, sid, name, name_en, role) in enumerate(chunks):
        end = chunks[i + 1][0] if i + 1 < len(chunks) else len(text)
        body = text[pos:end]
        st = re.search(r"stats:\s*\{\s*hp:\s*([\d.]+)\s*,\s*atk:\s*([\d.]+)\s*,\s*def:\s*([\d.]+)\s*,\s*speed:\s*([\d.]+)", body)
        stats = {"hp": 100, "atk": 15, "def": 4, "speed": 10}
        if st:
            stats = {"hp": float(st.group(1)), "atk": float(st.group(2)), "def": float(st.group(3)), "speed": float(st.group(4))}
        test = bool(re.search(r"testBuild:\s*true", body[:800]))
        abs_ = extract_A_blocks(body)
        specs.append({
            "id": sid, "name": name, "nameEn": name_en, "role": role,
            "stats": stats, "test": test, "abilities": abs_,
        })
    return specs


def class_meta(text):
    cid = re.search(r"id:\s*'([a-z]+)'\s*,\s*\n\s*name:\s*'([^']+)'", text)
    res = re.search(r"resource:\s*\{\s*type:\s*'([^']+)'\s*,\s*name:\s*'([^']+)'", text)
    sec = re.search(r"secondary:\s*\{\s*type:\s*'([^']+)'\s*,\s*name:\s*'([^']+)'", text)
    return {
        "id": cid.group(1) if cid else "?",
        "name": cid.group(2) if cid else "?",
        "res": (res.group(1), res.group(2)) if res else ("", ""),
        "sec": (sec.group(1), sec.group(2)) if sec else ("", ""),
    }


def parse_mastery():
    t = STATE.read_text(encoding="utf-8")
    out = {}
    for m in re.finditer(
        r"([a-z]+)_([a-z_]+):\s*\{\s*name:\s*'([^']+)'\s*,\s*effect:\s*'([^']+)'\s*,\s*kind:\s*'([^']+)'\s*,\s*pctAt120:\s*([\d.]+)",
        t,
    ):
        out[f"{m.group(1)}:{m.group(2)}"] = {
            "name": m.group(3), "effect": m.group(4), "kind": m.group(5), "pct": float(m.group(6)),
        }
    return out


def load_all():
    classes = []
    for p in sorted(BAL.glob("*-abilities.js")):
        if p.name == "apply-all.js":
            continue
        text = p.read_text(encoding="utf-8")
        meta = class_meta(text)
        specs = split_specs(text)
        # engineer / warrior class id from filename if parse missed
        if meta["id"] == "?" or meta["id"] in ("stagger", "swd", "chaos_bolt"):
            fname = p.stem.replace("-abilities", "")
            meta["id"] = fname
            meta["name"] = CLASS_RU.get(fname, fname)
        classes.append({"meta": meta, "specs": specs, "file": p.name})
    # keep known order
    order = ["warrior", "paladin", "hunter", "rogue", "priest", "deathknight",
             "shaman", "mage", "warlock", "monk", "druid", "engineer"]
    classes.sort(key=lambda c: order.index(c["meta"]["id"]) if c["meta"]["id"] in order else 99)
    return classes


def fmt_num(v):
    if v is None:
        return ""
    if isinstance(v, float) and v == int(v):
        return str(int(v))
    if isinstance(v, float):
        return str(round(v, 2)).rstrip("0").rstrip(".")
    return str(v)


def cost_line(ab):
    bits = []
    if ab.get("c") is not None:
        bits.append(f"c:{fmt_num(ab['c'])}")
    if ab.get("g"):
        bits.append(f"g:{fmt_num(ab['g'])}")
    if ab.get("cs"):
        bits.append(f"cs:{fmt_num(ab['cs'])}")
    if ab.get("gs"):
        bits.append(f"gs:{fmt_num(ab['gs'])}")
    r = ab.get("r")
    if isinstance(r, dict):
        parts = []
        for k, lab in (("b", "кровь"), ("f", "лёд"), ("u", "нечесть"), ("any", "любая")):
            if r.get(k):
                parts.append(f"{lab}×{r[k]}")
        if parts:
            bits.append("руны " + "+".join(parts))
    if ab.get("rp"):
        bits.append(f"+{fmt_num(ab['rp'])} RP")
    return " ".join(bits) or "0"


def flat_of(ab):
    if ab.get("fl") is not None:
        return ab["fl"]
    if ab.get("flat") is not None:
        return ab["flat"]
    return None


def est_raw(flat, atk, scale):
    if flat is None:
        return None
    return round(atk * scale * (float(flat) / FLAT_REF), 1)


def engine_dot(ab):
    aid = ab.get("id") or ""
    t = str(ab.get("t") or "")
    ad = ab.get("applyDot") if isinstance(ab.get("applyDot"), dict) else None
    if aid in DOT_FORCE or t == "dot":
        tick = flat_of(ab)
        return f"форс 3р · тик от flat {fmt_num(tick)}" + (f" (файл {fmt_num(ad.get('flat'))}т×{ad.get('turns')})" if ad else "")
    if ad:
        return f"applyDot жив: {fmt_num(ad.get('flat'))}т×{ad.get('turns')}"
    return ""


def owner_of(cid, sid):
    return "ты" if f"{cid}:{sid}" in OWNER_YOU else "ИИ / тест"


# ── review notes ──
# Предложения — ТЗ, цифры не вшиты. Колонка «Задача мне» пустая.

ENGINE_ROWS = [
    ["finisher_cp", "Финишер серии", "abilityDamageRaw берёт flat, локальный power×CP не используется. 1 CP = 5 CP.",
     "Множить сырой урон (или flat) от потраченных очков. Подсказка уже врёт в плюс.",
     "Разбойник все, ферал, пала WoG/заря если combo — нет, у пала ES. В основном рога/друид."],
    ["dot_applyDot", "type:dot / DOT_ABILITY_IDS", "Все чистые доты 3 раунда. applyDot.turns и applyDot.flat игнор. Тик = большой flat кнопки.",
     "Если есть applyDot — тик и длина оттуда; flat кнопки = первый хлопок. Нет applyDot — оставить 3р.",
     "Шаман FS, аффли, шадоу, гаррота, лунки, липкая бомба…"],
    ["buff_power1", "Баф без atkMod", "A() ставит power:1. Пустой баф = +100% ATK.",
     "Дефолт убрать. +атаки только при явном atkMod.",
     "Demo Dark Soul без atkMod, чай/TFT если забыли флаг."],
    ["dark_soul_pet", "Dark Soul и вечные петы", "petTurnsLeft == null → ставит 3. Основной демон пропадает.",
     "Вечным (null) только +ATK / не писать таймер.", "Демонология"],
    ["holy_shock_crit", "Шок небес", "Крит в ветке heal и ещё раз в dealDmg.",
     "Один ролл, в dealDmg skipCrit если уже крутили.", "Пал Свет"],
    ["fury_stacks", "Иск. Неистовства", "Любой cost==0 сбрасывает стаки: кик, жажда, FA.",
     "Сбрасывать только урон-скилл без ярости, не кик/FA/BT.", "Воин Неистовство"],
    ["applyStatus_id", "applyStatus затирает id", "Стаки печати / CS не копятся.",
     "Стаки плюсовать, не заменять бафф с тем же id.", "Пал Защита CS"],
    ["tempHp_leak", "Last Stand / VB", "maxHP растёт, в новый пак бафф снимают без отката пула.",
     "Снимать tempHp при конце боя / старте пака.", "Вар Защита, ДК Кровь"],
    ["felstorm_twice", "Felstorm", "Скилл бьёт в switch и ещё раз пет-AoE после.",
     "Один удар: либо кнопка, либо пет.", "Демонология"],
    ["ibf_id", "IBF / hasMajorDef", "Ищет id icebound, бафф может быть другим id.",
     "Вешать id icebound и учить holster оба.", "ДК все"],
    ["kick_ui", "Кик UI", "interruptPrimary у щита паладина не всегда в подсветке каста.",
     "Подсветка = type interrupt ИЛИ interruptPrimary.", "Пал Защита"],
    ["atonement", "Disc искупление", "lifesteal в кастера, не в раненого союзника. Pain Supp на себя.",
     "Хил самого низкого %HP союзника. DR на клик.", "Жрец Послушание"],
    ["loot_atk", "Лут +атака", "Крутит мёртвый power, на flat не садится.",
     "Множить getEff.atk или abilityDamageRaw.", "все flat-киты"],
    ["mark_loot", "Метка охотника в луте", "Текст +40% урона, код −def.",
     "Либо текст, либо код. Не врать.", "Охотник"],
    ["pet_damage_loot", "Шмот pet_damage", "Может не читаться в бою.",
     "Проверить getEff пета.", "Хант, лок, инженер"],
    ["armor_floor1", "Пол 1 урон", "raw − def×0.5/0.12, затем max(1). Мелкий физ = 1.",
     "Не трогать в первом патче (поедет ключ). Пометить в тултипе «до брони».", "все физ ST"],
    ["cd1_not_bug", "cd:1", "Тик в начале хода → кнопка каждый твой ход. Так задумано.",
     "НЕ менять семантику глобально. Если нужен пропуск — писать cd:2.", "хант KC, лава, keg…"],
    ["missing_mw", "Оружие волн / фульминация / FS→лава", "В lite нет. Описания у эле/энх честные.",
     "Если спек должен от этого играть — дописать 1 механику, не все сразу.", "Шаман эле/энх"],
    ["missing_diseases", "Болезни ДК как два id", "Частично applyDot, SS/Festering их не читают.",
     "Два id (озноб/чума), SS/Festering смотрят наличие.", "ДК Лёд/Нечесть"],
    ["havoc_cleave", "Havoc", "Сейчас −10% урона цели, не метка клеава.",
     "Метка: след. ST дублируется на помеченного.", "Разрушение"],
]

ROLE_ROWS = [
    ["танк", "atk≈12, speed 8 (хмель 10 ок), стена FA 2х КД≥8, авария HP всегда FA, танк-финишер не лучший ДД пака, FA-урон не лучшее AoE.",
     "Пал SotR 80+30 и Ardent 60%/3/КД6 сильнее стены вара. Хмель Guard %HP без КД. ДК atk 15. Blood blade 20% парир.",
     "Выровнять holster. Вар = блок/парир+авто-Реванш только с парира. Пал = AS-кик + земля/Суд. ДК = DS-хил + IBF. Хмель = шат→очистка→уклон. Страж = bleed-пак + кик."],
    ["хил", "реген 5–7, AoE без КД не тащит весь пак в одиночку, СТ без маны не полтанка, кик не всем (шаман + ткач — да).",
     "Слово славы 80 / 3ES / 0 маны. Цепь 40т на 5. Disc не Disc (искупление в себя).",
     "WoG сейв, не каждый пулл. Цепь скачет и слабеет. Disc бьёт и хилит рейда. Святой/ресто/друид/ткач — разные глаголы."],
    ["дд", "Ни один no-CD ST не жирнее ядра. 5 CP заметно > 1 CP. 1 клев + 1 ST + 1 кик на слот. FA-окно не ×1.6 два баффа.",
     "Flux 30т c:0 + луч 64. Demo DS без atkMod/+пет. Arms Slam 30 без КД. Combat KS 18×2 всем. Barrage 18×2.",
     "Сначала движок (CP, дот, баф). Потом потолки. Каждому спеку 1 глагол, от которого крутится ротация."],
]

# spec key -> review
SPECS = {}


def S(key, **kw):
    SPECS[key] = kw


S("warrior:arms",
  play="Колосс (−броня) → Смертельный/Слам + кровотечение 4р. Вихрь КД9 даёт размах → Героический клевит.",
  want="Тот же глагол: окно Колосса + bleed. Слам не должен быть лучше Казни в казни и лучше МС всегда.",
  engine="Размах уже есть. Иск. bleed жив.",
  add="—",
  cut="Слам 30 без КД — либо КД 1, либо ниже МС. Вихрь КД 9 мёртв или резать КД и 1 хит. Казнь > двух сламов в окне ≤35%.",
  risk="Высокий ST через Слам. Твой кит — не ломать Колосс/МС/bleed.")
S("warrior:fury",
  play="Жажда (хил, cd1, без flat → power 1.05) + Выпад 18×2. Иск. стаки с ярости, кик сбрасывает.",
  want="Набирать ярость → дамп Выпад/Казнь. Жажда топливо, не выключатель иск.",
  engine="Починить сброс стаков. Жажда без flat бьёт как 105% атаки — честно подписать или дать flat.",
  add="—",
  cut="Не стакать два окна. Вихрь как клев, не второй ST.",
  risk="Иск. сейчас враг кика.")
S("warrior:protection",
  play="Щит/Реванш/Гром. Блок+парир, авто-Реванш с парира. Стена FA КД12. LS ест ход, tempHp течёт.",
  want="Активный блок, ярость с ударов, ответ-Реванш. Магия ест в лицо — ок как ниша.",
  engine="tempHp. LS FA + свой id.",
  add="—",
  cut="Гром FA 20т — либо не FA, либо слабее (FA-урон ≠ лучшее AoE). Демо-крик дорогой.",
  risk="Цельный танк. Не давать авто-Реванш с блока.")
S("paladin:holy",
  play="Шок (+ES, хил или урон, двойной крит) → Свет/Вспышка. WoG 80 за 3ES без маны. Заря 30/2ES.",
  want="Дамп ES в сейв, Шок генератор, мана на большие хилы. Без кика (контракт).",
  engine="Один крит шока. WoG не combo-scale.",
  add="Маяк? только если сам попросишь — жирный.",
  cut="WoG вниз (сейв). Шок урон не 2.25×. Серый kick убрать с хила.",
  risk="Самый жирный СТ-сейв в игре.")
S("paladin:protection",
  play="AS кик+щит, Суд×Освящение, SotR 80+30, HotR 43/3ES как ST, Ardent 60%/3/6, CS стаки мёртвые.",
  want="AS = кик+клев+щит. Земля тикает. SotR короткий сейв-удар. HotR ген+клев, не второй казнь.",
  engine="UI кика, стаки CS, серый kick вынуть.",
  add="—",
  cut="SotR ≤ танк-финишер. Ardent ближе к стене вара. HotR ген.",
  risk="Лучший танк-holster сейчас.")
S("paladin:retribution",
  play="CS/Суд ген ES → Вердикт 38 / Буря 40 за 4ES. Молот FA execute power 1.45. Инквизиция +15% 2х.",
  want="На 1 цели Вердикт > Буря. Буря пак. Гнев окно. Иск. Длань Света чувствуется (сейчас 13%).",
  engine="—",
  add="Инквизиция может жрать ES (MoP).",
  cut="Буря 3ES и слабее Вердикта на 1 цели ИЛИ оставить 4/40 но тогда Вердикт жирнее.",
  risk="Твой кит. Не ломать Гнев-крылья без нужды.")
S("hunter:beast_mastery",
  play="Kill Command cd1 fl26 каждый ход, бьёт даже труп пета. Гнев+Rapid оба FA. Кика нет.",
  want="Пет — основной урон. KC каденс, живой пет. Один кик на класс (не всем трём).",
  engine="getMainPet живой. pet_damage лут.",
  add="Counter Shot только MM (слот кик-дд).",
  cut="KC cd:2. Не стакать два полных окна.",
  risk="Пак ломает ключ окном.")
S("hunter:marksmanship",
  play="Barrage 18×2 всем лучше Aimed. Химера слабее Aimed. Укус висит мёртвым.",
  want="Aimed dump, Химера сигнатура на КД, Barrage пак ниже Aimed на 1 цели.",
  engine="—",
  add="Кик только этому спеку.",
  cut="Barrage hits или цифру. Укус убрать или баф под Химеру.",
  risk="Пак-король.")
S("hunter:survival",
  play="Explosive cd1 24+дот каждый ход. Ловушка и BA отдельные.",
  want="Заряды разрывного / пропуск хода. BA бафает ES. Ловушка дебафф пака.",
  engine="type:dot у sticky-подобных — ES в списке? explosive id не в DOT_FORCE, applyDot жив если type damage. Сейчас t:dot? смотри реестр.",
  add="—",
  cut="ES не автоатака.",
  risk="Как KC.")
S("rogue:assassination",
  play="Мясорубка ген CP, Envenom/Rupture финишеры без скейла. Garrote type dot.",
  want="Яды + Envenom растёт от CP. Garrote удар+дот+тишина-окно.",
  engine="CP. Garrote вынуть из DOT_FORCE, type damage+applyDot.",
  add="Яд на веере.",
  cut="—",
  risk="Пока 1=5 CP баланс бессмысленен.")
S("rogue:combat",
  play="SS ген, Evis без скейла, Killing Spree 18×2 всем бесплатно, Blade Flurry +12% ATK не клев.",
  want="Клев (BF = %ST по остальным). Evis от CP. KS не 36т/цель.",
  engine="CP. BF клев.",
  add="Кик свой уже есть (пинок).",
  cut="KS в ST-окно или 1 хит AoE.",
  risk="Пак-король.")
S("rogue:subtlety",
  play="Ambush всегда. Dance окно слабое. Evis/Rupture без CP.",
  want="Ambush только в Dance. Find Weakness. CP-скейл.",
  engine="CP. Условие Dance на Ambush.",
  add="Find Weakness (+урон в окне).",
  cut="—",
  risk="Без окна саб = хуже боя.")
S("priest:discipline",
  play="Щит + большие хилы. Кара вампирит в себя 55%. Священный огонь в DOT_FORCE. Pain Supp на себя.",
  want="Бьёшь врага → хил раненого. Щит prevent. PS внешний сейв.",
  engine="Atonement в союзника. holy_fire вынуть из списка. PS на клик.",
  add="Искупление как правило, не lifesteal.",
  cut="Greater как чужой кит — можно Solace (урон+мана).",
  risk="Без движка Disc не Disc.")
S("priest:holy",
  play="Набор хилов + Круг КД2 + Слово 42. Играбельно. Кика нет (и не давать).",
  want="Смарт-AoE (Круг/PoH), Обновление на раненых, Слово сейв.",
  engine="—",
  add="Гимн / PoM позже.",
  cut="Кара не нужна как лицо.",
  risk="Низкий.")
S("priest:shadow",
  play="Доты все 3р. Сферы → чума. Кика нет. Шип мёртвый vs Flay.",
  want="Держать SWP+VT, Blast сферы, чума дамп. Кик (Silence).",
  engine="applyDot длины. Кик.",
  add="Silence КД 2–3. Mind Sear пак.",
  cut="Шип → кик или Sear.",
  risk="Слот кик-дд без кика.")
S("deathknight:blood",
  play="DS 24 +25% вампир, Heart/Boil руна крови, Coil 26 за RP как болт, Bone 40, atk 15, парир 20%, IBF id, VB FA + tempHp.",
  want="Танк atk 12. Дамп RP в рунический удар (физ+угроза), не нечестивый болт. DS ответ. IBF видит holster.",
  engine="IBF id, tempHp, atk.",
  add="Болезни как у остальных ДК (сейчас кровь без чумы).",
  cut="Парир 20% → ~10%. Coil → RS.",
  risk="Жирный holster + вампир.")
S("deathknight:frost",
  play="Oblit 30 не смотрит болезни. Howling клев+озноб. FS RP-дамп. Руны крови никто не тратит.",
  want="Болезни → Oblit жирнее. Кровь куда-то (конверсия или plague). Howling пак.",
  engine="Два id болезней. Бонус Oblit если оба.",
  add="Руна крови = лёд на Oblit/Howling или отдельный plague.",
  cut="—",
  risk="1/3 ресурса фейковая.")
S("deathknight:unholy",
  play="SS/Festering не продлевают. Coil 28. Outbreak КД8 AoE+дот. DT/вурдалак, горгулья. Нет IBF. Festering не читает доты.",
  want="Болезни держать, Festering продлевает, SS растёт от них, Coil дамп RP, пет лицо.",
  engine="Продление дотов. IBF. cost stun/rez если появятся — не 15 «рун» в cost.",
  add="IBF как у льда. Два id болезней.",
  cut="DT не 100% аптайм.",
  risk="Ближе всех ДК к фантазии, дыры в глаголе.")
S("shaman:elemental",
  play="Лава 32 КД1 ядро. Молния filler. FS type:dot 3р. ES простой урон. Цепь AoE. Blast слабее Лавы. Нет крита по FS.",
  want="FS сетап → Лава награда (крит или +flat). ES дамп (фульминация-lite: стаки со щита). Цепь пак.",
  engine="FS→Лава. Опционально стаки щита на ES.",
  add="1 механика: Лава видит FS. Не обязательно полный MW.",
  cut="Blast либо жирнее Лавы на КД, либо талант-кнопка.",
  risk="Без FS Лава = просто большая кнопка.")
S("shaman:enhancement",
  play="SS 28 + LL 26 КД1. FS под нову. Unleash баф +15%. Nova 16 всегда. Волки. LB filler без MW.",
  want="Две руки: буря и лава. Nova только если FS на цели. Unleash слабее ядер.",
  engine="Nova требует FS. MW не обязателен в первом проходе (честно написано).",
  add="Проверка FS на Nova.",
  cut="Unleash не +15% на 3х если это сильнее SS.",
  risk="Две кнопки КД1 = автоатаки.")
S("shaman:restoration",
  play="Быстрина 22+HoT (applyHot 5р жив). Цепь 40 на 5 −5%/скачок. Волна 30 > всплеск 24. Ливень HoT. Тотем summon. FS 3р. Link 15 +DR +выравнивание каждый удар.",
  want="Цепь лицо AoE. Быстрина HoT-якорь. Всплеск авария > волны. Link рейд-КД, не каждый хит выравнивать 5 раз.",
  engine="FS applyDot. Тотем потока тики если хочешь «как написано».",
  add="Кик уже есть (класс). Unleash +20% след. хилы — проверить что healAmp жив (у ресто есть).",
  cut="Цепь вниз или сильнее decay. Link не каждый удар (раз в раунд).",
  risk="Твой кит. Цепь+Link тащат ключ в паре с палом.")
S("mage:arcane",
  play="Blast/ракеты/барраж. Иск. «магистр маны» = просто +урон, не от % маны.",
  want="Стаки аркана → барраж дамп. Мана важна. Кик (антимагия) есть у мага как класс.",
  engine="—",
  add="Иск. от текущей маны или стаков.",
  cut="—",
  risk="Средний.")
S("mage:fire",
  play="Пиро 36/12 всегда. Живая бомба type:dot 3р. Комбуст.",
  want="Пиро награда за глыбу/прок, не вечный нук. Бомба тикает свою длину.",
  engine="applyDot. Прок Hot Streak-lite: след. пиро дешевле/FA.",
  add="1 прок: крит → пиро FA или дешевле.",
  cut="База пиро вниз, в проке вверх.",
  risk="Вечный пиро = нет цикла.")
S("mage:frost",
  play="Стрела, копьё 18 всегда < стрелы, шар, пурга, кольцо.",
  want="Копьё dump по заморозке. Шар/пурга пак. Кольцо контроль+урон.",
  engine="Метка freeze / shatter: копьё 28 по cc.",
  add="Shatter-lite.",
  cut="База копья вниз, по freeze вверх.",
  risk="Без shatter копьё мёртвое.")
S("warlock:affliction",
  play="3 дота все 3р (файл 5–6). Haunt 32/осколок. Drain ген. Dark Soul +25% ок (есть atkMod).",
  want="Повесить 3 дота разных длин → Haunt/Malefic по дотам жирнее.",
  engine="applyDot длины. Haunt +% если 3 дота.",
  add="Бонус от числа дотов.",
  cut="—",
  risk="Без длин доты = три одинаковые кнопки.")
S("warlock:demonology",
  play="SB power 1.15, SF 1.7/осколок, HoG AoE, Meta +30%, Dark Soul без atkMod (+100% баг), Felstorm дважды, бесы/страж.",
  want="Петы лицо. Meta окно. SF спендер. HoG бесы. DS не убивает стража и не +100%.",
  engine="DS таймер, DS atkMod явный, Felstorm 1 раз. SF на flat.",
  add="—",
  cut="Подписать DS. SF flat 34–36 уровня Haunt.",
  risk="Твой кит + баги движка.")
S("warlock:destruction",
  play="Испепел/поджиг ген осколок. Хаос 40. Immolate 3р. Havoc −10% урона цели (не клев). Shadowburn execute.",
  want="Доты+испепел → Хаос. Havoc = клев в метку. Ливень пак.",
  engine="Havoc метка. Immolate длина.",
  add="Havoc клев.",
  cut="—",
  risk="Хаос 40 ок как сигнатура если цикл жив.")
S("monk:brewmaster",
  play="Keg 40 КД1 AoE, BoK 45, Guard 45% HP без КД, уклон+стаки без капа, speed 10, шат 25% FA.",
  want="Шат→очистка→уклон. Keg клев. Guard щит flat+КД. Кап lucky.",
  engine="Кап стаков уклона. Guard не %HP.",
  add="—",
  cut="Guard flat ~40 КД2. BoK не 45 если keg 40. Лотос платный.",
  risk="Сильный танк с дырами. Твой кит — шат оставить.")
S("monk:mistweaver",
  play="Туманы HoT + подъём AoE. Чай +20% хилы. Revival 34 КД5. Кик нужно дать (контракт хил-кик).",
  want="Сад HoT, подъём по туманам, чай окно, кик.",
  engine="Чай healAmp проверить (есть). Не power:1.",
  add="Кик (как у шамана).",
  cut="—",
  risk="Играбелен если чай не +100%.")
S("monk:windwalker",
  play="Jab 40 эн → RSK 30 / BoK 28 / кулаки AoE. ToD execute. Чи-финишеры в списке FINISHER но secondary chi не combo — скейл не тот.",
  want="Чи спендеры. Лапа сетап (−броня), не лучший урон. Кулаки пак.",
  engine="Не применять combo-скейл к ци. Лапа дебафф.",
  add="Tiger Palm −броня цели.",
  cut="Лапа не 16т лучший filler если спендеры жирнее — ок. Следить чай если общий id.",
  risk="Ци как язык.")
S("druid:balance",
  play="Луна/солнце доты 3р, звёзды, ураган. Затмение в engineNeeds.",
  want="Доты держать, затмение меняет школу/кнопку, ураган пак.",
  engine="Затмение-lite: чередование или бафф после дота.",
  add="1 бафф затмения.",
  cut="—",
  risk="Без затмения = две одинаковые кнопки.")
S("druid:feral",
  play="Rake/Rip type:dot 3р. CP-финишер ferocious без скейла. Берсерк окно.",
  want="Bleed-пак: rake держать, rip от CP, укус дамп.",
  engine="CP. Rip/Rake длины из applyDot. Вынуть из форса или читать applyDot.",
  add="—",
  cut="—",
  risk="Как рога.")
S("druid:guardian",
  play="Mangle/Thrash/Lacerate. Шкура. Growl FA. Может быть озон-блок (общий танк).",
  want="Bleed-пак + шкура + Skull Bash кик. Не блок-воин.",
  engine="Убрать озон у стража (уже исключён? нет — исключены brew/пал/кровь). Страж имеет озон+15% блок — не его фантазия.",
  add="Skull Bash кик. Убрать озон.",
  cut="Mangle не без КД если это единственный ген.",
  risk="Чужой holster.")
S("druid:restoration",
  play="Reju/regrowth/lifebloom HoT (applyHot жив если не HOT_SPELLS-only). WG. Железный лай.",
  want="HoT-ramp, WG пак, кора сейв. Без кика.",
  engine="HOT_SPELLS форсит 3р если нет applyHot — проверить реестр.",
  add="Железный лай если нет.",
  cut="WG не 190 на 5 без КД.",
  risk="Связка с цепью шамана — не два короля AoE.")
S("engineer:mechanist",
  play="Гаечный/заклёпки/плазма ST. Турель и ходун за детали. Нет AoE лица. Кик свой.",
  want="Пет/турель ST. Личный урон filler. Турель клевит.",
  engine="Пет жив для турели.",
  add="Турель авто-клев.",
  cut="—",
  risk="0 AoE — ниша или дыра.")
S("engineer:sapper",
  play="Липкая type:dot 3р (файл 5×4). Кассеты/ракеты/подрыв 32. Remote 26 голый ST.",
  want="Липкая сетап → remote жрёт её и взрывает. Подрыв пак-КД.",
  engine="applyDot. Remote читает бомбу на цели.",
  add="Детонатор.",
  cut="Подрыв не каждый пак бесплатно 32+иск.",
  risk="Пак-король сапёра.")
S("engineer:tinkerer",
  play="Zap 13 +пар+деталь. Flux c:0 g:5 gs:2 fl:30 (не бесплатно без пара — ген). Луч 64 КД4. Курица/разрушитель. Отладка ST↔AoE. Воскрешатель жрёт пета.",
  want="Гаджеты+пет. Flux не жирнее луча без КД. Луч КД-нук. Отладка режим пета.",
  engine="Иск. pet_tune шанс, не 0% гаджетов.",
  add="Иск. бафает zap/flux/ray чуть-чуть или шанс гения.",
  cut="Flux вниз или цена пара. Луч оставить нуком.",
  risk="Твой кит. Потолок урона ключа. Цифры писать честно: 30т вес, +5 пар, +2 детали.")


BUTTON_PROP = {
    # key class:spec:id -> (propose, why)
    "warrior:arms:slam": ("КД 1 или ~26т, ниже МС+bleed", "Слам не лицо спека"),
    "warrior:arms:whirlwind": ("КД 6, 1 хит ~11 или оставить 2 хита слабее", "КД 9 мёртв; 18 пак не король"),
    "warrior:arms:execute": ("В окне ≤35% жирнее двух сламов", "Иначе кнопка мертва"),
    "warrior:fury:bt": ("Не сбрасывать иск.; подписать урон (нет flat)", "Топливо"),
    "warrior:protection:last_stand": ("FA + свой id + снять tempHp", "Авария не ест ход / не течёт"),
    "warrior:protection:thunder": ("Не FA или ≤13т", "FA-урон ≠ лучшее AoE"),
    "paladin:holy:word_glory": ("Сейв ~50–55т, не 80 каждый пулл", "Потолок СТ без маны"),
    "paladin:holy:holy_shock": ("Один крит", "Стоп 2.25×"),
    "paladin:protection:sot_r": ("Удар вниз, броню оставить", "Танк-финишер"),
    "paladin:protection:hot_r": ("Ген + клев, не 43 ST", "Вернуть молот-ген"),
    "paladin:protection:ardent": ("Ближе к стене вара (−50/2/8)", "Holster"),
    "paladin:protection:avengers": ("UI кика; bounce чуть выше; +ES?", "Лицо танка"),
    "paladin:retribution:divine_storm": ("На 1 цели < Вердикта", "Иначе всегда буря"),
    "shaman:restoration:ch": ("Сильнее decay или ниже база", "Не лучший хил во всём"),
    "shaman:restoration:chw": ("Авария ≥ волны", "Иначе кнопка мертва"),
    "shaman:restoration:flame_shock": ("Читать applyDot (4×6 или как решишь)", "Сейчас 7т×3"),
    "shaman:restoration:spirit_link": ("Выравнивание 1 раз/раунд", "Не 5 раз за 5 ударов"),
    "shaman:elemental:lv": ("Бонус/крит если FS на цели", "Сетап"),
    "shaman:enhancement:fire_nova": ("Без FS на цели — резать/не бить", "Текст = код"),
    "warlock:demonology:dark_soul": ("atkMod явный, не трогать вечных петов", "Баг +100% / таймер"),
    "warlock:demonology:felstorm": ("Один удар", "Двойной AoE"),
    "warlock:demonology:soul_fire": ("flat вместо power 1.7", "Честный вес"),
    "engineer:tinkerer:flux_bolt": ("Ниже 30 или цена пара; в тексте не «бесплатно»", "c:0 g:5 gs:2 fl:30"),
    "engineer:tinkerer:death_ray": ("Оставить нук КД, проверить DoT в тексте", "64т КД4"),
    "monk:brewmaster:guard": ("flat щит + КД, не 45% HP", "Стена каждый ход"),
    "priest:discipline:smite": ("Искупление в раненого, не в себя", "Фантазия Disc"),
    "priest:discipline:holy_fire": ("Вынуть из DOT_FORCE", "Удар+дот+искупление"),
    "priest:discipline:pain_supp": ("На клик, не на себя", "Внешний сейв"),
    "rogue:assassination:envenom": ("Скейл от CP", "Движок"),
    "rogue:combat:killing_spree": ("Не 18×2 всем", "Пак-король"),
    "rogue:combat:eviscerate": ("Скейл от CP", "Движок"),
    "hunter:beast_mastery:kill_cmd": ("КД 2, живой пет", "Не автоатака по трупу"),
    "hunter:marksmanship:barrage": ("Ниже Aimed на 1 цели", "Пак ≠ ST"),
    "mage:fire:pyroblast": ("База ниже; в проке награда", "Цикл"),
    "mage:frost:ice_lance": ("Слабее стрелы; по freeze жирнее", "Shatter"),
    "deathknight:blood:death_coil": ("Заменить на рунический удар / угроза", "Не болт"),
    "deathknight:unholy:festering": ("Продлевать болезни", "Глагол"),
}


def now_line(ab):
    bits = [ab.get("t") or "?"]
    fl = flat_of(ab)
    if fl is not None:
        bits.append(f"{fmt_num(fl)}т")
        if ab.get("hits"):
            bits.append(f"×{ab['hits']}")
    elif ab.get("p") not in (None, 1):
        bits.append(f"power {fmt_num(ab.get('p'))}")
    elif ab.get("t") == "buff" and not ab.get("atkMod") and not ab.get("dr") and not ab.get("cm"):
        bits.append("баф без atkMod (риск +100%)")
    bits.append(cost_line(ab))
    if ab.get("cd"):
        bits.append(f"КД {fmt_num(ab['cd'])}")
    if ab.get("fa") or ab.get("freeAction"):
        bits.append("FA")
    ed = engine_dot(ab)
    if ed:
        bits.append(ed)
    if ab.get("d"):
        bits.append("«" + str(ab["d"])[:80] + "»")
    return " · ".join(str(x) for x in bits if x)


def style_header(ws, row, n):
    for c in range(1, n + 1):
        cell = ws.cell(row, c)
        cell.fill = fill(HEAD)
        cell.font = FONT_H
        cell.alignment = Alignment(wrap_text=True, vertical="center")
        cell.border = THIN


def style_row(ws, row, n, bg=None):
    for c in range(1, n + 1):
        cell = ws.cell(row, c)
        cell.font = FONT
        cell.alignment = WRAP
        cell.border = THIN
        if bg:
            cell.fill = fill(bg)
    # last col comment always yellow
    ws.cell(row, n).fill = fill(YEL)


def set_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def title(ws, text, sub, cols):
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=cols)
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=cols)
    a = ws.cell(1, 1, text)
    a.font = FONT_T
    a.alignment = WRAP
    b = ws.cell(2, 1, sub)
    b.font = FONT_S
    b.alignment = WRAP
    ws.row_dimensions[1].height = 22
    ws.row_dimensions[2].height = 48
    ws.freeze_panes = "A4"
    ws.sheet_view.showGridLines = False
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.auto_filter.ref = None
    ws.sheet_properties.pageSetUpPr.fitToPage = True


def put(ws, r, vals):
    for i, v in enumerate(vals, 1):
        ws.cell(r, i, v if v is not None else "")


def finish_sheet(ws, header_row, n_cols, last_row, comment_col):
    if last_row > header_row:
        ws.auto_filter.ref = f"A{header_row}:{get_column_letter(n_cols)}{last_row}"
    ws.auto_filter.ref = ws.auto_filter.ref
    ws.freeze_panes = f"A{header_row + 1}"
    for r in range(header_row + 1, last_row + 1):
        ws.row_dimensions[r].height = 48
    # data validation empty - user types freely
    ws.freeze_panes = f"A{header_row + 1}"


def build():
    classes = load_all()
    mastery = parse_mastery()
    wb = Workbook()

    # ── Читай ──
    ws = wb.active
    ws.title = "Читай"
    title(ws, "Mythic Key — полный обзор баланса (Тест)",
          "12 августа 2026. Цифры НЕ вшиты в игру. Оценка +8 / +15 / среднее — до брони, без шмота, вес т × атака спека × скейл ключа. "
          "Жёлтая колонка «Задача мне» — пиши сюда поручения. Потом пришли файл — сделаю только помеченное.",
          6)
    ws.merge_cells("A4:F20")
    how = (
        "КАК ПОЛЬЗОВАТЬСЯ\n"
        "1) Лист «Кнопки» — факты из class-balance. Фильтр по классу/спеку. Сюда не спорь с текстом «бесплатно 30» — смотри колонки c/g/flat.\n"
        "2) Лист «Спеки» — от чего играет сейчас и что предлагаю. Одна строка = спек.\n"
        "3) Листы классов — те же кнопки + «Предлагаю / Зачем». Пустое предложение = не горит, можно дописать сам.\n"
        "4) «Движок» — сначала это, иначе цифры врут (доты 3р, CP, баф +100%).\n"
        "5) «Пассивки» — искусность и классовые пассивки (блок, парир, петы).\n\n"
        "КОЛОНКА «Задача мне» (жёлтая): пиши конкретно. Примеры:\n"
        "  • сделай\n"
        "  • сделай, но WoG 60 не 50\n"
        "  • не трогать\n"
        "  • сначала движок, цифры потом\n"
        "  • выкинуть кнопку, заменить на X\n\n"
        "Источник «ты» = кит, который ты собирал (вар, пал, ресто, демон, хмель, изобретатель). "
        "«ИИ / тест» = остальное, включая всего ДК.\n"
        "cd:1 в этой игре = каждый твой ход (не баг). Популярные ярлыки не перевожу: ST, AoE, DoT, HoT, CD, FA, CP, ES, RP, кик.\n"
        "Формула: сырой = atk_спека × скейл × (flat / 15). Скейл +8 = 1.09, +15 = 1.195. Броня потом режет физ сильнее магии."
    )
    cell = ws.cell(4, 1, how)
    cell.alignment = WRAP
    cell.font = FONT
    ws.row_dimensions[4].height = 280
    set_widths(ws, [22, 22, 22, 22, 22, 22])

    # ── Движок ──
    ws = wb.create_sheet("Движок")
    cols = ["id", "Что", "Сейчас в коде", "Предлагаю", "Кого задевает", "Задача мне"]
    title(ws, "Движок — факты и чинки",
          "Без этого ребаланс т — лотерея. Не менять cd:1 глобально.", len(cols))
    put(ws, 4, cols)
    style_header(ws, 4, len(cols))
    r = 5
    for row in ENGINE_ROWS:
        put(ws, r, list(row) + [""])
        bg = ENG
        if row[0] == "cd1_not_bug":
            bg = FACT
        style_row(ws, r, len(cols), bg)
        r += 1
    set_widths(ws, [18, 22, 44, 44, 36, 36])
    finish_sheet(ws, 4, len(cols), r - 1, 6)

    # ── Роли ──
    ws = wb.create_sheet("Роли")
    cols = ["Роль", "Контракт", "Сейчас ломает", "Предлагаю", "Задача мне"]
    title(ws, "Контракты ролей", "Чтобы патч не разъехался. holster ≠ winrate.", len(cols))
    put(ws, 4, cols)
    style_header(ws, 4, len(cols))
    r = 5
    for row in ROLE_ROWS:
        put(ws, r, list(row) + [""])
        style_row(ws, r, len(cols), PROP)
        r += 1
    set_widths(ws, [10, 50, 50, 50, 36])
    finish_sheet(ws, 4, len(cols), r - 1, 5)

    # ── Спеки ──
    ws = wb.create_sheet("Спеки")
    cols = ["Класс", "Спек", "Роль", "Источник", "hp/atk/def/spd", "Искусность",
            "Сейчас играет от", "Предлагаю играть от", "Движок / добавить", "Кнопки сменить", "Риск", "Задача мне"]
    title(ws, "36 специализаций — глагол",
          "Сначала прочитай «Движок». Потом галочки сюда и в жёлтый столбец.", len(cols))
    put(ws, 4, cols)
    style_header(ws, 4, len(cols))
    r = 5
    for cl in classes:
        cid = cl["meta"]["id"]
        for sp in cl["specs"]:
            key = f"{cid}:{sp['id']}"
            note = SPECS.get(key, {})
            m = mastery.get(key, {})
            mast = f"{m.get('name', '—')} · {m.get('kind', '')} {fmt_num(m.get('pct', ''))}%@120"
            if m.get("effect"):
                mast += " · " + m["effect"]
            st = sp["stats"]
            stat = f"{fmt_num(st['hp'])}/{fmt_num(st['atk'])}/{fmt_num(st['def'])}/{fmt_num(st['speed'])}"
            put(ws, r, [
                CLASS_RU.get(cid, cid), sp["name"], ROLE_RU.get(sp["role"], sp["role"]),
                owner_of(cid, sp["id"]), stat, mast,
                note.get("play", "см. кнопки"),
                note.get("want", ""),
                (note.get("engine") or "") + ((" | add: " + note["add"]) if note.get("add") and note.get("add") != "—" else ""),
                note.get("cut", ""),
                note.get("risk", ""),
                "",
            ])
            style_row(ws, r, len(cols), P0 if owner_of(cid, sp["id"]) == "ты" else FACT)
            r += 1
    set_widths(ws, [16, 18, 8, 12, 16, 36, 42, 42, 36, 40, 28, 36])
    finish_sheet(ws, 4, len(cols), r - 1, 12)

    # ── Кнопки registry ──
    ws = wb.create_sheet("Кнопки")
    cols = ["Класс", "Спек", "Роль", "Источник", "id", "Имя", "Тип", "Цена/ген", "КД", "FA",
            "Вес т", "hits", "power", "Школа", "DoT/HoT файл", "DoT факт движка",
            "оценка +8", "оценка +15", "средн 8/15", "atk спека", "Задача мне"]
    title(ws, "Реестр всех кнопок (факты)",
          f"Скейл +8={SCALE8:.3f}, +15={SCALE15:.3f}. Оценка = atk×скейл×flat/15, в «т», до брони и без шмота. "
          "Кик/диспел/пурга инжектятся в ui.js — их может не быть в этом списке.",
          len(cols))
    put(ws, 4, cols)
    style_header(ws, 4, len(cols))
    r = 5
    all_abs = []
    for cl in classes:
        cid = cl["meta"]["id"]
        for sp in cl["specs"]:
            atk = sp["stats"]["atk"]
            for ab in sp["abilities"]:
                fl = flat_of(ab)
                ad = ab.get("applyDot") if isinstance(ab.get("applyDot"), dict) else None
                ah = ab.get("applyHot") if isinstance(ab.get("applyHot"), dict) else None
                per = ""
                if ad:
                    per = f"DoT {fmt_num(ad.get('flat'))}т×{ad.get('turns')}"
                if ah:
                    per = (per + " | " if per else "") + f"HoT {fmt_num(ah.get('flat'))}т×{ah.get('turns')}"
                e8 = est_raw(fl, atk, SCALE8)
                e15 = est_raw(fl, atk, SCALE15)
                eavg = est_raw(fl, atk, AVG_SC)
                row = [
                    CLASS_RU.get(cid, cid), sp["name"], ROLE_RU.get(sp["role"], sp["role"]),
                    owner_of(cid, sp["id"]), ab.get("id") or "", ab.get("n") or "",
                    ab.get("t") or "", cost_line(ab), fmt_num(ab.get("cd") or 0),
                    "да" if (ab.get("fa") or ab.get("freeAction")) else "",
                    fmt_num(fl) if fl is not None else "",
                    fmt_num(ab.get("hits")) if ab.get("hits") else "",
                    fmt_num(ab.get("p")) if ab.get("p") not in (None, 1) else "",
                    ab.get("school") or "",
                    per, engine_dot(ab),
                    e8 if e8 is not None else "",
                    e15 if e15 is not None else "",
                    eavg if eavg is not None else "",
                    fmt_num(atk),
                    "",
                ]
                put(ws, r, row)
                style_row(ws, r, len(cols), FACT)
                all_abs.append((cid, sp, ab))
                r += 1
    set_widths(ws, [16, 16, 8, 12, 16, 26, 12, 22, 6, 5, 8, 6, 8, 10, 22, 28, 10, 10, 10, 8, 32])
    finish_sheet(ws, 4, len(cols), r - 1, 21)

    # ── Пассивки ──
    ws = wb.create_sheet("Пассивки")
    cols = ["Класс", "Спек", "Что", "Имя", "Суть", "Предлагаю поиграть", "Задача мне"]
    title(ws, "Искусность и пассивки",
          "Иск. при рейтинге 120. Пассивки из passives.js — блок/парир/петы тоже входят в урон и танк.",
          len(cols))
    put(ws, 4, cols)
    style_header(ws, 4, len(cols))
    r = 5
    PASSIVE_EXTRA = [
        ("warrior", "protection", "пассивка", "Щит с озона", "+15% блок (общий танк)", "Оставить — ниша вара"),
        ("warrior", "protection", "пассивка", "Одной левой", "+7% парир, авто-Реванш", "Только с парира, не с блока"),
        ("warrior", "arms", "пассивка", "Кровотечение", "МС/Колосс/Героический 4р", "Жив, не форсить 3р"),
        ("paladin", "все", "пассивка", "Добродетель", "25% вернуть каждую ES", "Ок"),
        ("paladin", "protection", "пассивка", "Святой щит / Защитник", "+15% блок +10% броня", "Ок"),
        ("paladin", "protection", "иск.", "Божественный оплот", "Только AS, 80%@120", "Жирно; 50%?"),
        ("paladin", "retribution", "иск.", "Длань Света", "13%@120 на holy", "Слабо, весь урон уже light"),
        ("paladin", "holy", "иск.", "Озарённое исцеление", "Эхо % хила 2х", "Ок"),
        ("deathknight", "все", "пассивка", "Рунный цикл", "6 рун, +5 RP/ход", "RP между пуллами не обнулять в 0"),
        ("deathknight", "blood", "пассивка", "Кровяной клинок", "+20% парир полный иммун", "Много → ~10%"),
        ("deathknight", "unholy", "пассивка", "Воскрешение мертвеца", "Постоянный вурдалак", "Ок"),
        ("shaman", "restoration", "иск.", "Глубокие воды", "+хил по потерянному HP, 20%@120 полная ≤30%", "Ок, лицо ресто"),
        ("warlock", "demonology", "иск.", "Мастер-демонолог", "Только петы 10%@120", "Мало если петы лицо"),
        ("engineer", "tinkerer", "пассивка", "Ходячая жестянка", "+1 деталь / 2 хода", "Ок"),
        ("engineer", "tinkerer", "иск.", "Гений инженерии", "Шанс прокачки пета 12%@120, 0% гаджетам", "Шанс + чуть zap/flux"),
        ("monk", "brewmaster", "пассивка", "Пьяный задира + ещё повезёт", "+6% уклон, стаки без капа", "Кап стаков 3, dodge≤60%"),
        ("druid", "guardian", "пассивка", "Щит с озона", "Страж получает общий блок 15%", "Убрать — не его ниша"),
    ]
    for row in PASSIVE_EXTRA:
        cid, sid, kind, name, effect, prop = row
        put(ws, r, [CLASS_RU.get(cid, cid), sid, kind, name, effect, prop, ""])
        style_row(ws, r, len(cols), PROP)
        r += 1
    for cl in classes:
        cid = cl["meta"]["id"]
        for sp in cl["specs"]:
            key = f"{cid}:{sp['id']}"
            m = mastery.get(key)
            if not m:
                continue
            # skip if already in extra by name
            put(ws, r, [CLASS_RU.get(cid, cid), sp["name"], "иск. @120", m["name"],
                        f"{m['kind']} {fmt_num(m['pct'])}% · {m['effect']}", "", ""])
            style_row(ws, r, len(cols), FACT)
            r += 1
    set_widths(ws, [16, 16, 14, 26, 56, 40, 36])
    finish_sheet(ws, 4, len(cols), r - 1, 7)

    # ── per class ──
    sheet_names = {
        "warrior": "Воин", "paladin": "Паладин", "hunter": "Охотник", "rogue": "Разбойник",
        "priest": "Жрец", "deathknight": "ДК", "shaman": "Шаман", "mage": "Маг",
        "warlock": "Чернокнижник", "monk": "Монах", "druid": "Друид", "engineer": "Инженер",
    }
    for cl in classes:
        cid = cl["meta"]["id"]
        ws = wb.create_sheet(sheet_names.get(cid, cid)[:31])
        cols = ["Спек", "id", "Имя", "Сейчас (факт)", "Предлагаю", "Зачем", "Задача мне"]
        meta = cl["meta"]
        res = f"рес. {meta['res'][1] or meta['res'][0]}"
        sec = f" · втор. {meta['sec'][1]}" if meta["sec"][0] else ""
        title(ws, f"{CLASS_RU.get(cid, cid)} — кнопки",
              res + sec + ". Жёлтый столбец — задания. Пустое «Предлагаю» = сам допиши или оставь.",
              len(cols))
        put(ws, 4, cols)
        style_header(ws, 4, len(cols))
        r = 5
        for sp in cl["specs"]:
            key = f"{cid}:{sp['id']}"
            note = SPECS.get(key, {})
            # spec banner row
            banner = (
                f"{sp['name']} · {ROLE_RU.get(sp['role'])} · {owner_of(cid, sp['id'])} · "
                f"играет: {note.get('play', '—')}"
            )
            put(ws, r, [sp["name"], "—спек—", banner, note.get("want", ""), note.get("cut", ""), note.get("engine", ""), ""])
            style_row(ws, r, len(cols), P1)
            r += 1
            for ab in sp["abilities"]:
                pk = f"{cid}:{sp['id']}:{ab.get('id')}"
                pr = BUTTON_PROP.get(pk, ("", ""))
                put(ws, r, [
                    sp["name"], ab.get("id") or "", ab.get("n") or "",
                    now_line(ab), pr[0], pr[1], "",
                ])
                style_row(ws, r, len(cols), FACT)
                r += 1
        set_widths(ws, [16, 16, 28, 56, 40, 32, 36])
        finish_sheet(ws, 4, len(cols), r - 1, 7)

    # print area / header
    for s in wb.worksheets:
        s.oddHeader.left.text = "Тест · полный обзор · не вшито"
        s.oddFooter.left.text = "Жёлтое = задача мне"
        s.oddFooter.right.text = "стр. &P / &N"
        s.sheet_view.zoomScale = 100

    wb.save(OUT)
    print("wrote", OUT, "sheets", len(wb.sheetnames))
    n_ab = sum(len(sp["abilities"]) for cl in classes for sp in cl["specs"])
    n_sp = sum(len(cl["specs"]) for cl in classes)
    print("classes", len(classes), "specs", n_sp, "abilities", n_ab)


if __name__ == "__main__":
    build()
