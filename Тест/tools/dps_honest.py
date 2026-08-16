# -*- coding: utf-8 -*-
"""
Честный урон 1 / 5 / 10: все спеки (танк и хил тоже) жмут макс. урон.

СТ — ротация в одну. 5 и 10 — ротация в пак. Хилы не лечат.
Доты не перекастывают, пока висят. Финишер с 5 очков серии.
"""
from __future__ import annotations

import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent))
from build_role_compare import (  # noqa: E402
    COMBO_MULT,
    CRIT_EXP,
    FINISHERS,
    FLAT_REF,
    HORIZON,
    PET_FLAT,
    SKIP_IDS,
    SKIP_TYPES,
    ab_cd,
    ab_flat,
    hit_t,
    is_finisher,
    load_kits,
    num,
    uses_combo,
)

# Исповедь в ките t=heal, но бьёт врага на 30т
DAMAGE_AS_HEAL = {"penance"}

# призыв → (т за ход пета, длительность ходов хозяина)
SUMMON_TICK = {
    "hellfiend": (34.0, 5),
    "raise_ghoul": (15.0, 2),
    "summon_water": (40.0 / 3.0, 3),
    "xuen": (18.0, 3),
    "shadowfiend": (18.0, 4),
    "fire_ele": (19.0, 4),
    "dire": (17.0, 3),
}

ROLE_TAG = {"tank": "ТАНК", "healer": "ХИЛ", "dps": "ДД"}


def usable_dmg(ab):
    if not ab or not ab.get("id"):
        return False
    if ab.get("id") in SKIP_IDS:
        return False
    t = ab.get("t") or ""
    if ab.get("id") in DAMAGE_AS_HEAL:
        return True
    if t in SKIP_TYPES:
        return False
    if t in ("heal", "heal_aoe", "shield", "taunt"):
        return False
    return True


def is_atk_buff(ab):
    return bool(ab.get("atkMod") or ab.get("am")) and (ab.get("t") or "") == "buff"


def is_def_buff(ab):
    t = ab.get("t") or ""
    if t == "buff" and (ab.get("dr") or ab.get("dmgReduce")) and not ab.get("atkMod"):
        return True
    return False


def splash_t(ab, atk, n):
    extra = 0.0
    if n > 1 and ab.get("splashFlat"):
        extra += atk * num(ab["splashFlat"]) / FLAT_REF * (n - 1)
    if n > 1 and ab.get("cleaveFlat"):
        extra += atk * num(ab["cleaveFlat"]) / FLAT_REF * min(2, n - 1)
    return extra


def combo_scale(ab, combo):
    if not is_finisher(ab):
        return 1.0
    n = min(5, max(0, int(combo)))
    return COMBO_MULT.get(n, COMBO_MULT[1]) if n else COMBO_MULT[1]


def simulate_dps(kit, n_targets):
    atk0 = kit["stats"]["atk"]
    rtype = kit["rtype"]
    regen = kit["regen"]
    res = 20.0 if rtype == "rage" else (6.0 if rtype == "runes" else 100.0)
    res_max = 6.0 if rtype == "runes" else 100.0
    if kit["classId"] == "paladin":
        sec_max, sec = 5.0, 3.0
    elif kit["classId"] == "monk" and kit["specId"] != "mistweaver":
        sec_max, sec = 5.0, 0.0
    elif kit["classId"] == "warlock":
        sec_max, sec = 4.0, 1.0
    elif kit["classId"] == "priest" and kit["specId"] == "shadow":
        sec_max, sec = 3.0, 0.0
    else:
        sec_max, sec = 5.0, 0.0
    combo = 0
    cds = {}
    charges = {}
    dots = {}
    log = []
    dmg_st = 0.0
    dmg_sum = 0.0
    atk_mod = 1.0
    atk_mod_left = 0
    summons = []  # {left, tick}
    snake_left = 0

    abs_ = [a for a in kit["abilities"] if usable_dmg(a)]
    for a in abs_:
        cds[a["id"]] = 0
        if a.get("ch"):
            charges[a["id"]] = int(num(a["ch"]))

    pet_stand = PET_FLAT.get(kit["key"], 0.0)

    def can_cast(a):
        if cds.get(a["id"], 0) > 0 and not (a.get("ch") and charges.get(a["id"], 0) > 0):
            return False
        cs = num(a.get("cs"))
        if a.get("id") == "divine_storm" and cs > 3:
            cs = 3.0
        if cs and not uses_combo(kit) and sec + 1e-6 < cs:
            return False
        cost = num(a.get("c"))
        if rtype == "runes":
            if a.get("id") in ("fs",):
                return sec >= 35
            if a.get("id") in ("death_coil", "death_strike"):
                return sec >= 40
            need = 2.0 if a.get("id") in ("obliterate", "festering") else 1.0
            return res + 1e-6 >= need
        return res + 1e-6 >= cost

    def pay(a):
        nonlocal res, sec
        if rtype == "runes":
            if a.get("id") == "fs":
                sec -= 35
            elif a.get("id") in ("death_coil", "death_strike"):
                sec -= 40
            elif a.get("id") in ("obliterate", "festering"):
                res -= 2
            else:
                res -= 1
                sec = min(100.0, sec + 10)
        else:
            res -= num(a.get("c"))
        cs = num(a.get("cs"))
        if a.get("id") == "divine_storm" and cs > 3:
            cs = 3.0
        if cs and not uses_combo(kit):
            sec = max(0.0, sec - cs)
        if a.get("gs") and not uses_combo(kit):
            sec = min(sec_max, sec + num(a.get("gs")))
        if a.get("g"):
            res = min(res_max, res + num(a["g"]))
        if a.get("ch") and charges.get(a["id"], 0) > 0:
            charges[a["id"]] -= 1
            if charges[a["id"]] < int(num(a["ch"])) and cds.get(a["id"], 0) <= 0:
                cds[a["id"]] = ab_cd(a)
        elif ab_cd(a):
            cds[a["id"]] = ab_cd(a)

    def tick_dots():
        total = 0.0
        dead = []
        for k, d in dots.items():
            total += d["tick"] * d["n"]
            d["left"] -= 1
            if d["left"] <= 0:
                dead.append(k)
        for k in dead:
            del dots[k]
        return total

    def apply_dot(a, atk_now, used_combo):
        ad = a.get("applyDot") or {}
        fl = num(ad.get("flat"))
        turns = int(num(ad.get("turns"), 0))
        if (a.get("t") == "dot") and not fl:
            fl = ab_flat(a) or 0
            turns = turns or 3
        if fl <= 0 or turns <= 0:
            return
        tick = atk_now * fl / FLAT_REF
        if is_finisher(a) and used_combo > 0:
            tick *= COMBO_MULT.get(min(5, used_combo), 1.0) / COMBO_MULT[5]
        t = a.get("t") or ""
        tn = n_targets if t == "aoe" or a.get("applyDotAoe") else 1
        key = str(ad.get("id") or a.get("id"))
        dots[key] = {"tick": tick, "left": turns, "n": tn}

    def missing_dot_value(a, atk_now, left):
        ad = a.get("applyDot") or {}
        fl = num(ad.get("flat"))
        turns = int(num(ad.get("turns"), 0))
        if (a.get("t") == "dot") and not fl:
            fl = ab_flat(a) or 0
            turns = turns or 3
        if fl <= 0 or turns <= 0:
            return 0.0
        key = str(ad.get("id") or a.get("id"))
        cur = dots.get(key)
        if cur and cur["left"] > 1:
            return 0.0
        tick = atk_now * fl / FLAT_REF
        t = a.get("t") or ""
        tn = n_targets if t == "aoe" or a.get("applyDotAoe") else 1
        return tick * min(turns, left) * tn

    def score(a, atk_now, left, turn):
        t = a.get("t") or ""
        if is_def_buff(a):
            return -1.0
        if t == "buff" and not is_atk_buff(a) and a.get("id") not in SUMMON_TICK and a.get("id") != "jade_serpent":
            if t == "buff":
                return -1.0
        base = hit_t(a, atk_now)
        if a.get("id") == "penance":
            base = atk_now * 30.0 / FLAT_REF
            t = "damage"
        base *= combo_scale(a, combo)
        extra_dot = missing_dot_value(a, atk_now, left)
        splash = splash_t(a, atk_now, n_targets)
        if t == "aoe":
            dmg = base * n_targets + extra_dot + splash
        elif t == "dot":
            dmg = extra_dot + base * 0.05
        elif t == "summon":
            spec = SUMMON_TICK.get(a.get("id"))
            if a.get("id") == "jade_serpent":
                # 3т фокусу после хода каждого: хозяин + n врагов, 3 раунда
                dmg = 3.0 * 3 * (1 + n_targets)
            elif spec:
                dmg = spec[0] * spec[1]
            else:
                dmg = 20.0
        elif t == "buff":
            dmg = 6.0
        else:
            dmg = base + extra_dot + splash
        if n_targets == 1 and t == "aoe":
            dmg *= 0.22
        if n_targets >= 5 and t == "aoe":
            dmg *= 1.25
        if extra_dot > 0 and (t == "dot" or a.get("applyDot")):
            dmg *= 1.12
        if is_finisher(a) and combo < 5 and turn < HORIZON - 2:
            return -1.0
        return dmg

    def add_dealt(st_part, sum_part):
        nonlocal dmg_st, dmg_sum
        dmg_st += st_part
        dmg_sum += sum_part

    for turn in range(HORIZON):
        if atk_mod_left > 0:
            atk_mod_left -= 1
            if atk_mod_left <= 0:
                atk_mod = 1.0
        for a in abs_:
            if cds.get(a["id"], 0) > 0:
                cds[a["id"]] -= 1
                if cds[a["id"]] <= 0 and a.get("ch") and charges.get(a["id"], 0) < int(num(a["ch"])):
                    charges[a["id"]] = charges.get(a["id"], 0) + 1
                    if charges[a["id"]] < int(num(a["ch"])):
                        cds[a["id"]] = ab_cd(a)
        if rtype == "runes":
            res = min(res_max, res + 2.0)
        else:
            res = min(res_max, res + regen)

        dt = tick_dots()
        add_dealt(dt if n_targets == 1 else 0.0, dt)

        still = []
        for s in summons:
            add_dealt(s["tick"] if n_targets == 1 else 0.0, s["tick"])
            s["left"] -= 1
            if s["left"] > 0:
                still.append(s)
        summons = still
        if snake_left > 0:
            # после хода хозяина + каждого врага
            chunk = 3.0 * (1 + n_targets)
            add_dealt(3.0, chunk)
            snake_left -= 1

        atk_now = atk0 * atk_mod
        left = HORIZON - turn

        for a in abs_:
            if not (a.get("fa") or a.get("freeAction")):
                continue
            if not can_cast(a):
                continue
            if is_def_buff(a):
                continue
            if is_atk_buff(a) or a.get("id") in SUMMON_TICK or a.get("id") == "jade_serpent" or (a.get("t") in ("damage", "aoe") and (a.get("fa") or a.get("freeAction"))):
                if a.get("t") in ("damage", "aoe") and not is_atk_buff(a) and a.get("id") not in SUMMON_TICK:
                    # FA-урон (Молот гнева и т.п.) — только если не казнь; уже в SKIP
                    pass
                pay(a)
                if is_atk_buff(a):
                    atk_mod = 1.0 + num(a.get("atkMod") or a.get("am"))
                    atk_mod_left = int(num(a.get("bt") or a.get("buffTurns"), 2))
                    atk_now = atk0 * atk_mod
                if a.get("id") == "jade_serpent":
                    snake_left = max(snake_left, 3)
                spec = SUMMON_TICK.get(a.get("id"))
                if spec and a.get("id") != "jade_serpent":
                    summons.append({"tick": spec[0], "left": spec[1]})
                if a.get("t") in ("damage", "aoe") and not is_atk_buff(a):
                    base = hit_t(a, atk_now)
                    if a.get("t") == "aoe":
                        add_dealt(base if n_targets == 1 else 0.0, base * n_targets)
                    else:
                        add_dealt(base, base + splash_t(a, atk_now, n_targets))
                    apply_dot(a, atk_now, combo)
                log.append(f"х{turn + 1} FA {a.get('n')}")

        best = None
        best_sc = 0.0
        for a in abs_:
            fa = bool(a.get("fa") or a.get("freeAction"))
            t = a.get("t") or ""
            if fa and t not in ("damage", "aoe", "dot") and a.get("id") not in DAMAGE_AS_HEAL:
                continue
            if not can_cast(a):
                continue
            if is_finisher(a):
                if combo < 1:
                    continue
                if combo < 5 and turn < HORIZON - 2:
                    continue
            sc = score(a, atk_now, left, turn)
            if sc > best_sc:
                best_sc = sc
                best = a
        if not best:
            log.append(f"х{turn + 1} простой")
            continue

        a = best
        t = a.get("t") or ""
        used_combo = combo
        pay(a)
        if uses_combo(kit) and a.get("gs") and not is_finisher(a):
            combo = min(5, combo + int(num(a.get("gs"))))
        if is_finisher(a):
            combo = 0

        base = hit_t(a, atk_now)
        if a.get("id") == "penance":
            base = atk_now * 30.0 / FLAT_REF
            t = "damage"
        if is_finisher(a):
            base *= combo_scale(a, used_combo)
        apply_dot(a, atk_now, used_combo if is_finisher(a) else 0)
        splash = splash_t(a, atk_now, n_targets)

        if t == "aoe":
            add_dealt(base if n_targets == 1 else 0.0, base * n_targets + splash)
        elif t == "dot":
            pass
        elif t == "summon":
            if a.get("id") == "jade_serpent":
                snake_left = max(snake_left, 3)
            spec = SUMMON_TICK.get(a.get("id"))
            if spec and a.get("id") != "jade_serpent":
                summons.append({"tick": spec[0], "left": spec[1]})
        elif t == "buff":
            if is_atk_buff(a):
                atk_mod = 1.0 + num(a.get("atkMod") or a.get("am"))
                atk_mod_left = int(num(a.get("bt") or a.get("buffTurns"), 2))
        else:
            add_dealt(base, base + splash)

        log.append(f"х{turn + 1} {a.get('n')}")

    if pet_stand:
        if kit["key"] == "engineer:mechanist" and n_targets >= 5:
            add_dealt(pet_stand * 8, pet_stand * 8 + pet_stand * 1.7 * n_targets * 4)
        else:
            add_dealt(pet_stand * HORIZON, pet_stand * HORIZON)

    if n_targets == 1:
        tot = max(dmg_st, dmg_sum)
        dmg_st = tot
        dmg_sum = tot

    dmg_st *= CRIT_EXP
    dmg_sum *= CRIT_EXP
    return {
        "dmg_st": round(dmg_st, 1),
        "dmg_aoe": round(dmg_sum, 1),
        "log": " → ".join(log[:16]),
        "atk": atk0,
    }


def collect_honest_dmg():
    kits = load_kits()
    rows = []
    for k in kits:
        r1 = simulate_dps(k, 1)
        r5 = simulate_dps(k, 5)
        r10 = simulate_dps(k, 10)
        st, a5, a10 = r1["dmg_st"], r5["dmg_aoe"], r10["dmg_aoe"]
        rows.append({
            "role": k["role"],
            "cls": k["className"],
            "spec": k["specName"],
            "st": st,
            "a5": a5,
            "a10": a10,
            "r5": round(a5 / st, 2) if st else 0,
            "r10": round(a10 / a5, 2) if a5 else 0,
            "log1": r1["log"],
            "log5": r5["log"],
            "log10": r10["log"],
            "atk": r1["atk"],
        })
    return rows


def print_lists(rows):
    def dump(title, key):
        print("=" * 72)
        print(title)
        print("=" * 72)
        s = sorted(rows, key=lambda x: -x[key])
        for i, x in enumerate(s, 1):
            tag = ROLE_TAG.get(x["role"], x["role"])
            print(
                f"{i:2d} [{tag:4}] {x['cls']} {x['spec']:22}  "
                f"СТ={x['st']:7.1f}  5={x['a5']:7.1f}  10={x['a10']:7.1f}  "
                f"5/СТ={x['r5']}  10/5={x['r10']}"
            )
        print()

    dump("ЧЕСТНЫЙ УРОН — 1 цель (макс СТ)", "st")
    dump("ЧЕСТНЫЙ УРОН — сумма 5", "a5")
    dump("ЧЕСТНЫЙ УРОН — сумма 10", "a10")


def write_honest_dmg_sheets(wb, rows):
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    ink = "1A1A1A"
    head = "2C2416"
    head_f = "F7F1E3"
    fills = {"tank": "D9E8F5", "healer": "DFF0D8", "dps": "F8E6C8"}
    thin = Border(
        left=Side(style="thin", color="D0C4A8"),
        right=Side(style="thin", color="D0C4A8"),
        top=Side(style="thin", color="D0C4A8"),
        bottom=Side(style="thin", color="D0C4A8"),
    )
    wrap = Alignment(wrap_text=True, vertical="top")
    font = Font(name="Calibri", size=10, color=ink)
    font_h = Font(name="Calibri", size=10, bold=True, color=head_f)

    def put_sheet(name, headers, data, widths):
        if name in wb.sheetnames:
            del wb[name]
        ws = wb.create_sheet(name)
        ws.append(headers)
        for i in range(1, len(headers) + 1):
            c = ws.cell(1, i)
            c.fill = PatternFill("solid", fgColor=head)
            c.font = font_h
            c.alignment = Alignment(wrap_text=True, vertical="center")
            c.border = thin
        for ri, row in enumerate(data, 2):
            role = row[0]
            for ci, v in enumerate(row, 1):
                cell = ws.cell(ri, ci, v)
                cell.font = font
                cell.alignment = wrap
                cell.border = thin
                hex_ = fills.get(role)
                if hex_:
                    cell.fill = PatternFill("solid", fgColor=hex_)
        for i, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{len(data) + 1}"
        ws.row_dimensions[1].height = 28
        return ws

    ranked = []
    for key, label in (("st", "СТ"), ("a5", "5"), ("a10", "10")):
        order = {id(x): i + 1 for i, x in enumerate(sorted(rows, key=lambda z: -z[key]))}
        for x in rows:
            x[f"rank_{key}"] = order[id(x)]

    table = []
    for x in sorted(rows, key=lambda z: z["rank_st"]):
        table.append([
            x["role"], x["cls"], x["spec"],
            x["st"], x["a5"], x["a10"], x["r5"], x["r10"],
            x["rank_st"], x["rank_a5"], x["rank_a10"],
            x["log1"], x["log5"],
        ])
    put_sheet(
        "Честный урон",
        [
            "Роль", "Класс", "Спек",
            "СТ 12х", "Сумма 5", "Сумма 10", "5 / СТ", "10 / 5",
            "Место СТ", "Место 5", "Место 10",
            "Ротация 1", "Ротация 5",
        ],
        table,
        [10, 16, 20, 12, 12, 12, 10, 10, 10, 10, 10, 56, 56],
    )
    rots = [[x["role"], x["cls"], x["spec"], x["log1"], x["log5"], x["log10"]] for x in rows]
    put_sheet(
        "Честный урон ротации",
        ["Роль", "Класс", "Спек", "1 цель", "5 целей", "10 целей"],
        rots,
        [10, 16, 20, 70, 70, 70],
    )


def write_into_workbook(path):
    from openpyxl import Workbook, load_workbook

    path = Path(path)
    rows = collect_honest_dmg()
    if path.exists():
        wb = load_workbook(path)
    else:
        wb = Workbook()
        if "Sheet" in wb.sheetnames:
            del wb["Sheet"]
    write_honest_dmg_sheets(wb, rows)
    wb.save(path)
    return rows, path


def main(argv=None):
    argv = list(sys.argv[1:] if argv is None else argv)
    quiet_rot = "--без-ротации" in argv
    out = Path(__file__).resolve().parents[1] / "сравнение_ролей.xlsx"
    rows, path = write_into_workbook(out)
    print_lists(rows)
    if not quiet_rot:
        print("=" * 72)
        print("РОТАЦИЯ СТ / 5")
        print("=" * 72)
        for x in sorted(rows, key=lambda z: (z["role"] != "dps", z["cls"], z["spec"])):
            tag = ROLE_TAG.get(x["role"], x["role"])
            print(f"[{tag}] {x['cls']} {x['spec']}  СТ={x['st']}  5={x['a5']}")
            print(f"   1: {x['log1']}")
            print(f"   5: {x['log5']}")
            print()
    print("лист «Честный урон» →", path)


if __name__ == "__main__":
    main()
