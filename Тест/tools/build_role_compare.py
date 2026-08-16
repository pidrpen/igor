# -*- coding: utf-8 -*-
"""
Глобальное сравнение ролей (Тест-киты).
12 своих ходов, сетка 1 / 5 / 10 равносильных целей.
Книжные «т» через atk / FLAT_REF=15. Не бой, а оценка ротации.
"""
from __future__ import annotations

import math
import re
import sys
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.chart import BarChart, Reference

ROOT = Path(__file__).resolve().parents[1]
OUT = ROOT / "сравнение_ролей.xlsx"
sys.path.insert(0, str(Path(__file__).resolve().parent))
from build_full_review import load_all  # noqa: E402
from tank_resist import simulate_tank_resist  # noqa: E402
from disc_atonement import simulate_disc_atonement  # noqa: E402
from healer_honest import simulate_healer_honest  # noqa: E402

FLAT_REF = 15.0
HORIZON = 12
CRIT_EXP = 1.0 + 0.18 * 0.5  # 18% крит ×1.5 → ожидание 1.09
ENEMY_ST = 18.0
ENEMY_SPLASH = 7.0
PARTY = 5
HEALER_REF_ST = 32.0  # «типичный» СТ-хил за ход партнёра
COMBO_MULT = {0: 1.0, 1: 0.22, 2: 0.42, 3: 0.68, 4: 1.05, 5: 1.55}
FINISHERS = {
    "dispatch", "eviscerate", "rupture", "rip", "ferocious",
}
SKIP_TYPES = {"interrupt", "taunt", "cc", "dispel", "purge", "cleanse"}
SKIP_IDS = {
    "kick", "shock_wrench", "growl", "taunt", "provoke", "pet_rez",
    "debug_mode", "touch_death", "execute", "kill_shot", "hot_w",
}

PET_FLAT = {
    "hunter:beast_mastery": 22.0,
    "hunter:marksmanship": 14.0,
    "hunter:survival": 14.0,
    "warlock:demonology": 18.0,
    "warlock:affliction": 10.0,
    "warlock:destruction": 10.0,
    "deathknight:unholy": 15.0,
    "deathknight:frost": 8.0,
    "engineer:mechanist": 25.0,
    "engineer:tinkerer": 16.0,
    "mage:frost": 40.0 / 3.0,  # элементаль 3р из 12, грубо
    "monk:mistweaver": 3.0,  # змея не каждый ход хозяина; занижено
    "priest:discipline": 0.0,
}

INK = "1A1A1A"
HEAD = "2C2416"
HEAD_F = "F7F1E3"
BG = "F7F1E3"
TANK_C = "D9E8F5"
HEAL_C = "DFF0D8"
DPS_C = "F8E6C8"
WARN = "FFF3B0"
BAD = "F4D6D0"
GOOD = "C5E0B4"
THIN = Border(
    left=Side(style="thin", color="D0C4A8"),
    right=Side(style="thin", color="D0C4A8"),
    top=Side(style="thin", color="D0C4A8"),
    bottom=Side(style="thin", color="D0C4A8"),
)
WRAP = Alignment(wrap_text=True, vertical="top")
FONT = Font(name="Calibri", size=10, color=INK)
FONT_B = Font(name="Calibri", size=10, bold=True, color=INK)
FONT_H = Font(name="Calibri", size=10, bold=True, color=HEAD_F)
FONT_T = Font(name="Calibri", size=16, bold=True, color=HEAD)


def fill(hex_):
    return PatternFill("solid", fgColor=hex_)


def num(v, d=0.0):
    if v is None or v is False:
        return d
    try:
        return float(v)
    except (TypeError, ValueError):
        return d


def ab_flat(ab):
    if ab.get("fl") is not None:
        return num(ab["fl"])
    if ab.get("flat") is not None:
        return num(ab["flat"])
    return None


def ab_cd(ab):
    return int(num(ab.get("cd"), 0))


def parse_regen(text, spec_id):
    """regen / start / type from spec override or class resource."""
    chunks = []
    pat = re.compile(
        r"id:\s*'([a-z_]+)'\s*,\s*(?:\n\s*)?name:",
        re.M,
    )
    hits = list(pat.finditer(text))
    body = text
    for i, m in enumerate(hits):
        if m.group(1) == spec_id:
            end = hits[i + 1].start() if i + 1 < len(hits) else len(text)
            body = text[m.start():end]
            break
    ov = re.search(
        r"resourceOverride:\s*\{[^}]*type:\s*'([^']+)'[^}]*regen:\s*([\d.]+)",
        body,
    )
    if ov:
        st = re.search(r"start:\s*([\d.]+)", body[body.find("resourceOverride"):body.find("resourceOverride") + 220])
        return ov.group(1), num(ov.group(2)), num(st.group(1), 100) if st else 100.0
    cls = re.search(
        r"resource:\s*\{[^}]*type:\s*'([^']+)'[^}]*regen:\s*([\d.]+)",
        text,
    )
    if cls:
        return cls.group(1), num(cls.group(2)), 100.0
    return "none", 0.0, 100.0


def load_kits():
    classes = load_all()
    out = []
    for c in classes:
        cid = c["meta"]["id"]
        if cid in ("cheat", "demonhunter"):
            continue
        text = (ROOT / "class-balance" / c["file"]).read_text(encoding="utf-8")
        for s in c["specs"]:
            rtype, regen, start = parse_regen(text, s["id"])
            out.append({
                "classId": cid,
                "className": c["meta"]["name"],
                "specId": s["id"],
                "specName": s["name"],
                "role": s["role"],
                "stats": s["stats"],
                "rtype": rtype,
                "regen": regen,
                "start": start,
                "abilities": s["abilities"],
                "key": f"{cid}:{s['id']}",
            })
    return out


def hit_t(ab, atk):
    fl = ab_flat(ab)
    hits = max(1, int(num(ab.get("hits"), 1)))
    if fl is not None:
        return atk * fl / FLAT_REF * hits
    p = ab.get("p")
    if p is None:
        return 0.0
    return atk * num(p) * hits


def armor_dr(defn):
    # физ. броня: def*1000 / (def*1000 + 20000)
    d = defn * 1000.0
    return min(0.75, d / (d + 20000.0))


def tank_extra_dr(key):
    """Средний доп. сейв за пулл (не пик стены)."""
    extra = {
        "warrior:protection": 0.12,   # блок/парир + крик −15% часть времени
        "paladin:protection": 0.10,   # сейвы + освящение
        "deathknight:blood": 0.18,    # удар смерти / костяной / кровь вампира усреднённо
        "monk:brewmaster": 0.22,      # пошатывание 35% + очистка, часть уходит в пул
        "druid:guardian": 0.14,       # кожа / инстинкты / броня дикой
    }
    return extra.get(key, 0.0)


def is_finisher(ab):
    return ab.get("id") in FINISHERS or (num(ab.get("cs")) == 1 and ab.get("id") not in ("envenom",))


def usable(ab):
    if not ab or not ab.get("id"):
        return False
    if ab.get("id") in SKIP_IDS:
        return False
    t = ab.get("t") or ""
    if t in SKIP_TYPES:
        return False
    return True


def apply_hot_window(ab, atk, left):
    hot = ab.get("applyHot") or {}
    fl = num(hot.get("flat"))
    turns = int(num(hot.get("turns"), 0))
    if fl <= 0 or turns <= 0:
        return 0.0
    ticks = min(turns, left)
    return atk * fl / FLAT_REF * ticks


def apply_dot_window(ab, atk, n, left, combo_pts=0):
    dot = ab.get("applyDot") or {}
    fl = num(dot.get("flat"))
    turns = int(num(dot.get("turns"), 0))
    if fl <= 0 or turns <= 0:
        # type:dot without applyDot: 3 ticks of half? skip if no applyDot
        if (ab.get("t") == "dot") and ab_flat(ab):
            fl = ab_flat(ab)
            turns = 3
        else:
            return 0.0
    ticks = min(turns, left)
    tick = atk * fl / FLAT_REF
    if is_finisher(ab) and combo_pts > 0:
        tick *= COMBO_MULT.get(min(5, combo_pts), 1.0) / COMBO_MULT[5]
    t = ab.get("t") or ""
    targets = n if t == "aoe" or ab.get("applyDotAoe") else 1
    return tick * ticks * targets


def score_ab(ab, atk, n, role, combo, left):
    """Ожидаемая польза за этот ход (без FA)."""
    t = ab.get("t") or ""
    fa = bool(ab.get("fa") or ab.get("freeAction"))
    base = hit_t(ab, atk)
    if is_finisher(ab) and combo > 0:
        base *= COMBO_MULT.get(min(5, combo), 1.0)
    elif is_finisher(ab):
        base *= COMBO_MULT[1]
    dot = apply_dot_window(ab, atk, n, left, combo if is_finisher(ab) else 0)
    hot = apply_hot_window(ab, atk, left)
    if t == "aoe":
        dmg = base * n + dot
        heal = 0.0
    elif t == "heal_aoe":
        decay = num(ab.get("chainDecay"), 0.0)
        heal = 0.0
        m = 1.0
        for i in range(min(PARTY, max(1, n if n >= 5 else PARTY))):
            heal += base * m
            m *= (1.0 - decay)
        heal += hot * min(PARTY, 5)
        dmg = 0.0
    elif t == "heal":
        heal = base + hot
        dmg = 0.0
    elif t == "shield":
        heal = base * 0.85
        dmg = 0.0
    elif t == "dot":
        dmg = (base * 0.15 if base else 0.0) + apply_dot_window(ab, atk, n, left, combo if is_finisher(ab) else 0)
        heal = 0.0
    elif t == "buff":
        dmg = 8.0 if fa else 4.0  # маленький вес, чтобы повесить окно
        heal = 0.0
    elif t == "summon":
        dmg = 12.0
        heal = 0.0
    else:
        dmg = base + dot
        heal = 0.0
    ls = num(ab.get("lifesteal"))
    if ls and dmg:
        heal += dmg * ls
    hfd = num(ab.get("healFromDealt"))
    if hfd and dmg:
        heal += dmg * hfd
    if role == "healer":
        return heal * 1.15 + dmg * 0.25
    if n >= 5 and t == "aoe":
        return dmg * 1.05 + heal * 0.2
    if n == 1 and t == "aoe":
        return dmg * 0.55 + heal * 0.2
    return dmg + heal * 0.25


def uses_combo(kit):
    return kit["classId"] == "rogue" or kit["specId"] == "feral"


def put_dot(store, key, tick, turns, targets):
    store[key] = {"tick": tick, "left": turns, "n": targets}


def tick_store(store):
    total = 0.0
    dead = []
    for k, d in store.items():
        total += d["tick"] * d["n"]
        d["left"] -= 1
        if d["left"] <= 0:
            dead.append(k)
    for k in dead:
        del store[k]
    return total


def simulate(kit, n_targets):
    atk0 = kit["stats"]["atk"]
    hp = kit["stats"]["hp"]
    deff = kit["stats"]["def"]
    role = kit["role"]
    rtype = kit["rtype"]
    regen = kit["regen"]
    res = 20.0 if rtype == "rage" else (6.0 if rtype == "runes" else 100.0)
    res_max = 6.0 if rtype == "runes" else 100.0
    sec_need = max((int(num(a.get("cs"))) for a in kit["abilities"]), default=0)
    sec_max = 5.0
    if kit["classId"] == "paladin":
        sec_max = float(max(3, sec_need))
        sec = 3.0  # каждый пулл 3 Энергии Света
    elif kit["classId"] == "monk" and kit["specId"] != "mistweaver":
        sec_max = 5.0
        sec = 0.0
    elif kit["classId"] == "warlock":
        sec_max = 4.0
        sec = 1.0
    else:
        sec = 0.0
    combo = 0
    cds = {}
    charges = {}
    log = []
    dmg_st = 0.0
    dmg_aoe_sum = 0.0
    heal_st = 0.0
    heal_aoe = 0.0
    self_heal = 0.0
    atk_mod = 1.0
    atk_mod_left = 0
    dots = {}
    hots = {}

    abs_ = [a for a in kit["abilities"] if usable(a)]
    for a in abs_:
        cds[a["id"]] = 0
        if a.get("ch"):
            charges[a["id"]] = int(num(a["ch"]))

    pet = PET_FLAT.get(kit["key"], 0.0)

    def can_cast(a):
        if cds.get(a["id"], 0) > 0 and not (a.get("ch") and charges.get(a["id"], 0) > 0):
            return False
        cs = num(a.get("cs"))
        if a.get("id") == "divine_storm" and cs > 3:
            cs = 3.0
        if cs and not uses_combo(kit) and sec + 1e-6 < cs:
            return False
        cost = num(a.get("c"))
        if rtype == "runes":
            if a.get("id") == "fs":
                return sec >= 35
            if a.get("id") == "death_coil":
                return sec >= 40
            if a.get("id") == "death_strike":
                return sec >= 40
            need = 2.0 if a.get("id") in ("obliterate", "festering") else 1.0
            return res + 1e-6 >= need
        return res + 1e-6 >= cost

    def pay(a):
        nonlocal res, sec
        if rtype == "runes":
            if a.get("id") == "fs":
                sec -= 35
            elif a.get("id") == "death_coil":
                sec -= 40
            elif a.get("id") == "death_strike":
                sec -= 40
            elif a.get("id") in ("obliterate", "festering"):
                res -= 2
            else:
                res -= 1
                sec = min(100.0, sec + 10)
        else:
            res -= num(a.get("c"))
        cs = num(a.get("cs"))
        if a.get("id") == "divine_storm" and cs > 3:
            cs = 3.0
        if cs and not uses_combo(kit):
            sec = max(0.0, sec - cs)
        if a.get("gs") and not uses_combo(kit):
            sec = min(sec_max, sec + num(a.get("gs")))
        if a.get("g"):
            res = min(res_max, res + num(a["g"]))
        if a.get("ch") and charges.get(a["id"], 0) > 0:
            charges[a["id"]] -= 1
            if charges[a["id"]] < int(num(a["ch"])) and cds.get(a["id"], 0) <= 0:
                cds[a["id"]] = ab_cd(a)
        elif ab_cd(a):
            cds[a["id"]] = ab_cd(a)

    def apply_periodics(a, atk_now, used_combo):
        ad = a.get("applyDot") or {}
        fl = num(ad.get("flat"))
        turns = int(num(ad.get("turns"), 0))
        if (a.get("t") == "dot") and not fl:
            fl = ab_flat(a) or 0
            turns = turns or 3
        if fl > 0 and turns > 0:
            tick = atk_now * fl / FLAT_REF
            if is_finisher(a) and used_combo > 0:
                tick *= COMBO_MULT.get(min(5, used_combo), 1.0) / COMBO_MULT[5]
            t = a.get("t") or ""
            tn = n_targets if t == "aoe" or a.get("applyDotAoe") else 1
            put_dot(dots, str(ad.get("id") or a.get("id")), tick, turns, tn)
        ah = a.get("applyHot") or {}
        hf = num(ah.get("flat"))
        ht = int(num(ah.get("turns"), 0))
        if hf > 0 and ht > 0:
            put_dot(hots, str(ah.get("name") or a.get("id")), atk_now * hf / FLAT_REF, ht, 1)

    for turn in range(HORIZON):
        if atk_mod_left > 0:
            atk_mod_left -= 1
            if atk_mod_left <= 0:
                atk_mod = 1.0
        for a in abs_:
            if cds.get(a["id"], 0) > 0:
                cds[a["id"]] -= 1
                if cds[a["id"]] <= 0 and a.get("ch") and charges.get(a["id"], 0) < int(num(a["ch"])):
                    charges[a["id"]] = charges.get(a["id"], 0) + 1
                    if charges[a["id"]] < int(num(a["ch"])):
                        cds[a["id"]] = ab_cd(a)
        if rtype == "runes":
            res = min(res_max, res + 2.0)
        else:
            res = min(res_max, res + regen)

        d_tick = tick_store(dots)
        h_tick = tick_store(hots)
        dmg_st += d_tick if n_targets == 1 else 0
        dmg_aoe_sum += d_tick
        heal_st += h_tick
        if role == "tank":
            self_heal += h_tick

        atk_now = atk0 * atk_mod

        for a in abs_:
            if not (a.get("fa") or a.get("freeAction")):
                continue
            if (a.get("t") or "") != "buff":
                continue
            if not can_cast(a):
                continue
            pay(a)
            if a.get("atkMod"):
                atk_mod = 1.0 + num(a["atkMod"])
                atk_mod_left = int(num(a.get("bt") or a.get("buffTurns"), 2))
                atk_now = atk0 * atk_mod
            log.append(f"х{turn+1} FA {a.get('n')}")

        best = None
        best_sc = -1.0
        for a in abs_:
            fa = bool(a.get("fa") or a.get("freeAction"))
            t = a.get("t") or ""
            if fa and t not in ("damage", "aoe", "heal", "heal_aoe"):
                continue
            if not can_cast(a):
                continue
            if is_finisher(a):
                if combo < 1:
                    continue
                if combo < 5 and turn < HORIZON - 2:
                    continue
            sc = score_ab(a, atk_now, n_targets, role, combo, HORIZON - turn)
            if role == "healer":
                if n_targets == 1 and t == "heal_aoe":
                    sc *= 0.12
                if n_targets == 1 and t in ("heal", "shield"):
                    sc *= 1.35
                if n_targets >= 5 and t == "heal_aoe":
                    sc *= 1.4
                if n_targets >= 5 and t == "heal":
                    sc *= 0.55
            else:
                if n_targets == 1 and t == "aoe":
                    sc *= 0.35
                if n_targets >= 5 and t == "aoe":
                    sc *= 1.35
            if sc > best_sc:
                best_sc = sc
                best = a
        if not best:
            log.append(f"х{turn+1} простой")
            continue

        a = best
        t = a.get("t") or ""
        used_combo = combo
        pay(a)
        if uses_combo(kit) and a.get("gs") and not is_finisher(a):
            combo = min(5, combo + int(num(a.get("gs"))))
        if is_finisher(a):
            combo = 0

        base = hit_t(a, atk_now)
        if is_finisher(a):
            base *= COMBO_MULT.get(min(5, max(1, used_combo)), 1.0)
        apply_periodics(a, atk_now, used_combo if is_finisher(a) else 0)

        if t == "aoe":
            chunk = base * n_targets
            dmg_aoe_sum += chunk
            if n_targets == 1:
                dmg_st += base
        elif t == "heal_aoe":
            decay = num(a.get("chainDecay"), 0.0)
            m = 1.0
            h = 0.0
            for _i in range(PARTY):
                h += base * m
                m *= (1.0 - decay)
            heal_aoe += h
        elif t == "heal":
            heal_st += base
            if role == "tank":
                self_heal += base
        elif t == "shield":
            heal_st += base * 0.85
            if role == "tank":
                self_heal += base * 0.85
        elif t == "dot":
            pass
        elif t in ("buff", "summon", "debuff"):
            pass
        else:
            dmg_st += base
            dmg_aoe_sum += base
            if n_targets > 1 and a.get("cleaveFlat"):
                extra = atk_now * num(a["cleaveFlat"]) / FLAT_REF * min(2, n_targets - 1)
                dmg_aoe_sum += extra

        dealt_now = base * (n_targets if t == "aoe" else 1)
        ls = num(a.get("lifesteal"))
        if ls:
            sh = dealt_now * ls
            self_heal += sh
            heal_st += sh
        hfd = num(a.get("healFromDealt"))
        if hfd:
            sh = dealt_now * hfd
            self_heal += sh
            heal_st += sh
        if a.get("id") == "death_strike":
            pct = 0.15 if kit["specId"] == "blood" else 0.10
            self_heal += hp * pct
            heal_st += hp * pct

        log.append(f"х{turn+1} {a.get('n')}")

    if pet:
        if kit["key"] == "engineer:mechanist" and n_targets >= 5:
            dmg_st += pet * 8
            dmg_aoe_sum += pet * 8 + pet * 1.7 * n_targets * 4
        else:
            dmg_st += pet * HORIZON
            dmg_aoe_sum += pet * HORIZON

    dmg_st *= CRIT_EXP
    dmg_aoe_sum *= CRIT_EXP
    heal_st *= CRIT_EXP
    heal_aoe *= CRIT_EXP
    self_heal *= CRIT_EXP

    if n_targets == 1:
        tot = max(dmg_st, dmg_aoe_sum)
        dmg_st = tot
        dmg_aoe_sum = tot

    return {
        "dmg_st": round(dmg_st, 1),
        "dmg_aoe": round(dmg_aoe_sum, 1),
        "heal_st": round(heal_st, 1),
        "heal_aoe": round(heal_aoe, 1),
        "self_heal": round(self_heal, 1),
        "log": " → ".join(log[:18]),
        "hp": hp,
        "atk": atk0,
        "def": deff,
        "dr": min(0.72, armor_dr(deff) + tank_extra_dr(kit["key"])),
    }


def weakness_text(kit, rows):
    role = kit["role"]
    r1 = rows[1]
    r5 = rows[5]
    r10 = rows[10]
    bits = []
    if role == "dps":
        if r1["dmg_st"] < 250:
            bits.append("Низкий урон в одну цель за 12 ходов.")
        if r5["dmg_aoe"] < r1["dmg_st"] * 1.6:
            bits.append("Слабо растёт в 5 целей (мало области / доты только в одну).")
        if r10["dmg_aoe"] / max(1, r5["dmg_aoe"]) < 1.25:
            bits.append("Потолок в 10 почти как в 5: область не масштабируется.")
        if kit["rtype"] in ("mana", "energy", "focus") and kit["regen"] <= 6:
            bits.append("Ресурс сухой — ротация может проседать без окна.")
        if kit["key"] in PET_FLAT:
            bits.append("Часть урона на питомце: если пет мёртв, цифра падает.")
        if "feral" in kit["key"] or kit["classId"] == "rogue":
            bits.append("Серия: слив раньше 5 очков сильно режет финишер.")
    elif role == "healer":
        if r1["heal_st"] < 280:
            bits.append("СТ-хил за пулл ниже среднего.")
        if r5["heal_aoe"] < r1["heal_st"] * 0.7:
            bits.append("Мало области: в паке слабее, чем в одну цель.")
        if kit["specId"] == "discipline":
            bits.append("Считает игру от Искупления (Кара / Священный огонь / Исповедь во врага / пет), не Великое исцеление. Без баффа на цели корм не лечит.")
        if kit["specId"] == "mistweaver":
            bits.append("Змея и хоты: пик не сразу, нужен разгон.")
        if kit["regen"] <= 5:
            bits.append("Мана садится — длинный пулл хуже короткого.")
    else:
        if r1["dr"] < 0.35:
            bits.append("Мало снижения входящего без стен.")
        inc10 = tank_incoming(r10["dr"], 10)
        if inc10 > kit["stats"]["hp"] * 8:
            bits.append("В 10 целях соло не стоит: входящий кратно числу ударов.")
        if r1["self_heal"] < 40:
            bits.append("Мало самохила — без партнёра-хила запас кончается быстрее.")
        if kit["specId"] == "brewmaster":
            bits.append("Пошатывание откладывает урон: если не чистить, пул взорвётся.")
        if kit["specId"] == "blood":
            bits.append("Самохил завязан на Удар смерти и силу рун — без них сухой.")
    if not bits:
        bits.append("Явной дыры в этой модели нет — смотри ранг на сетке.")
    return " ".join(bits)


def tank_incoming(dr, n):
    # все n бьют танка СТ каждый раунд
    raw = ENEMY_ST * n * HORIZON
    return raw * (1.0 - dr)


def style_header(ws, row, cols, color=HEAD):
    for i in range(1, cols + 1):
        c = ws.cell(row, i)
        c.fill = fill(color)
        c.font = FONT_H
        c.alignment = Alignment(wrap_text=True, vertical="center")
        c.border = THIN


def write_sheet(ws, headers, rows, col_w, fills=None):
    ws.append(headers)
    style_header(ws, 1, len(headers))
    for ri, row in enumerate(rows, 2):
        for ci, v in enumerate(row, 1):
            cell = ws.cell(ri, ci, v)
            cell.font = FONT
            cell.alignment = WRAP
            cell.border = THIN
            if fills and fills[ri - 2]:
                cell.fill = fill(fills[ri - 2])
    for i, w in enumerate(col_w, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{len(rows)+1}"
    ws.freeze_panes = "A2"
    ws.row_dimensions[1].height = 28


def main():
    kits = load_kits()
    # simulate all
    pack = []
    for kit in kits:
        by_n = {}
        for n in (1, 5, 10):
            if kit["key"] == "priest:discipline":
                by_n[n] = simulate_disc_atonement(kit, n)
            else:
                by_n[n] = simulate(kit, n)
        pack.append((kit, by_n))

    healers = [(k, r) for k, r in pack if k["role"] == "healer"]
    avg_healer_st = sum(r[1]["heal_st"] for _, r in healers) / max(1, len(healers))

    wb = Workbook()

    # ── Метод ──
    ws = wb.active
    ws.title = "Метод"
    ws["A1"] = "Сравнение ролей — как читать"
    ws["A1"].font = FONT_T
    ws.merge_cells("A1:B1")
    lines = [
        ("Окно", "12 своих ходов. Это короткий пулл, не ключ целиком."),
        ("Сетка", "1 / 5 / 10 равносильных целей. Враги бьют так же «по книге»: 18т в танка за свой ход каждый."),
        ("т", "Боевая формула: atk × flat / 15. У танка atk 12, поэтому его 18т на кнопке ≈ 14.4 в бою. В таблицах — уже после atk."),
        ("Крит", "Ожидание 18% × 1.5 = ×1.09 ко всему исходящему."),
        ("Ротация", "Жадный выбор: доступно по ресурсу и КД, область ценнее при 5+ целях, финишер серии не раньше 5 очков (кроме конца окна). Без хода (бафф) вешается в начале хода."),
        ("Доты / хоты", "Тики, которые успевают пройти до конца 12 ходов. Повторный каст не двойнит уже висящее — считается окно до конца пулла."),
        ("Питомец", "Отдельная добавка за 12 ходов (боевой бот 25т, звери и т.д.). Механист в 5+ целях: 4 хода модуля область +70%."),
        ("Танк соло", "Индекс = (HP + самохил) / входящий. Входящий = 18т × число врагов × 12 × (1 − броня − средний сейв). Все бьют танка."),
        ("Танк с хилом", "К числителю добавляется средний СТ-хил всех хилов за те же 12 ходов (сейчас ≈ {0:.0f}т). Это «есть хил в группе», не идеальный сейв."),
        ("Хил СТ", "Сумма лечения в одну постоянно раненую цель за 12 ходов."),
        ("Хил АоЕ", "Область на 5 союзников (цепь с decay). На листе 10 — та же ротация, область ×2 как грубый «рейд»."),
        ("ДД", "СТ — урон в фокус. 5 и 10 — сумма по всем целям (область + фокус + доты в основную)."),
        ("Не модель", "Промахи, контроль, кики, казни по окну HP, смерть пета, искусность выше базы, шмот, крит-окна мага. Это сравнение китов, не симулятор ключа."),
        ("Божественная буря", "В данных стоит 4 Энергии Света при потолке 3. В этой таблице расход как у Вердикта (3), иначе область Воздаяния никогда не нажимается."),
        ("Читер", "Не входит."),
        ("Файл китов", "igor-main/Тест/class-balance/* после заливки S01–S35."),
        ("Послушание", "Не жадный хил. Держит Искупление (щит 5р / Щит небес 5р всем / Молитва 3р) и кормит Карой, Священным огнём, Исповедью во врага и Исчадием ада. Каждый носитель — 55% нанесённого. СТ = 1 носитель. Область = 5 носителей."),
        ("Хил честный", "Таблицы ниже на листе «Хилы». Искусность, крит, хоты, Искупление, Выбор света, змея/эхо Ткача, цепь −5%, Энергия Света и Добродетель, Громовой чай, Высвободить жизнь, Прилив. СТ = союзник 0. Область = сумма по 5."),
        ("Танк старая таблица", "Сверху на листе «Танки»: вход = 18т×N×12×(1 − броня − зашитый сейв). Кнопки на индекс почти не влияют."),
        ("Танк честный сейв", "Таблицы ниже на том же листе. Танк жмёт стены / блок / самохил / броню. Входящий режется бронёй из защиты, стаками Удара воина Света и Щита света, стенами, блоком, париром, уклоном, щитом, пошатыванием, искусностью Крови и Стража. Блок/парир/уклон — ожидание, не один ролл."),
        ("Честный Соло", "(HP + временный пул + самохил) / вход в HP. Вход в HP — что сняли с полоски (удары + тик пошатывания). Больше 1 ≈ карман перекрыл окно."),
        ("Честный с хилом", "Отдельный прогон: каждый ход после ударов капает средний СТ-хил хилов / 12. Может спасти от смерти и сдвинуть Удар смерти / Неистовое восстановление."),
    ]
    ws["A3"] = "Правило"
    ws["B3"] = "Как считали"
    style_header(ws, 3, 2)
    for i, (a, b) in enumerate(lines, 4):
        ws.cell(i, 1, a).font = FONT_B
        ws.cell(i, 2, b.format(avg_healer_st)).font = FONT
        ws.cell(i, 1).alignment = WRAP
        ws.cell(i, 2).alignment = WRAP
        ws.cell(i, 1).border = THIN
        ws.cell(i, 2).border = THIN
        ws.row_dimensions[i].height = 36
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 110

    # ── Сводка ──
    ws = wb.create_sheet("Сводка")
    tanks = [(k, r) for k, r in pack if k["role"] == "tank"]
    dps = [(k, r) for k, r in pack if k["role"] == "dps"]

    resist = {}
    resist_h = {}
    heal_tick = avg_healer_st / HORIZON
    for k, _r in tanks:
        resist[k["key"]] = {}
        resist_h[k["key"]] = {}
        for n in (1, 5, 10):
            resist[k["key"]][n] = simulate_tank_resist(k, n, 0.0)
            resist_h[k["key"]][n] = simulate_tank_resist(k, n, heal_tick)

    def rank_desc(items, keyfn):
        return sorted(items, key=keyfn, reverse=True)

    rows = [["Роль", "Вопрос", "1 цель", "5 целей", "10 целей", "Комментарий"]]
    # tanks
    for n, label in ((1, "1"), (5, "5"), (10, "10")):
        pass
    t_solo = {}
    t_heal = {}
    for n in (1, 5, 10):
        scored = []
        for k, r in tanks:
            inc = tank_incoming(r[n]["dr"], n)
            solo = (k["stats"]["hp"] + r[n]["self_heal"]) / max(1.0, inc)
            withh = (k["stats"]["hp"] + r[n]["self_heal"] + avg_healer_st) / max(1.0, inc)
            scored.append((k, solo, withh, inc))
        t_solo[n] = rank_desc(scored, lambda x: x[1])
        t_heal[n] = rank_desc(scored, lambda x: x[2])

    def names(lst):
        return " → ".join(
            f"{i+1}. {x[0]['className']} {x[0]['specName']}" for i, x in enumerate(lst)
        )

    rows.append([
        "Танк", "Кто лучше держит соло",
        names(t_solo[1]), names(t_solo[5]), names(t_solo[10]),
        "Индекс (HP+самохил)/входящий. В 10 все проседают — смотри порядок, не «выживет».",
    ])
    rows.append([
        "Танк", "Кто лучше держит с хилом",
        names(t_heal[1]), names(t_heal[5]), names(t_heal[10]),
        f"К самохилу добавлен средний СТ-хил хилов ({avg_healer_st:.0f}т / 12 ходов).",
    ])

    def honest_names(n, store):
        scored = []
        for k, _r in tanks:
            s = store[k["key"]][n]
            scored.append((k, s["solo"]))
        return names(rank_desc(scored, lambda x: x[1]))

    rows.append([
        "Танк", "Честный сейв соло",
        honest_names(1, resist), honest_names(5, resist), honest_names(10, resist),
        "Жмёт стены/блок/самохил. Соло = (HP + временный пул + самохил) / вход в HP.",
    ])
    rows.append([
        "Танк", "Честный сейв с хилом",
        honest_names(1, resist_h), honest_names(5, resist_h), honest_names(10, resist_h),
        f"Отдельный прогон: +{heal_tick:.1f} т хила каждый ход (средний СТ хилов / 12).",
    ])

    h_st = rank_desc(healers, lambda x: x[1][1]["heal_st"])
    h_aoe = rank_desc(healers, lambda x: x[1][5]["heal_aoe"])
    rows.append([
        "Хил", "Кто лучше лечит СТ",
        names([(k, r[1]["heal_st"], 0, 0) for k, r in h_st]),
        "—", "—",
        "Одна постоянно раненая цель. Область на этом листе не главная.",
    ])
    rows.append([
        "Хил", "Кто лучше лечит область",
        "—",
        names([(k, r[5]["heal_aoe"], 0, 0) for k, r in h_aoe]),
        names([(k, r[10]["heal_aoe"] * 2, 0, 0) for k, r in rank_desc(healers, lambda x: x[1][10]["heal_aoe"])]),
        "5 союзников. Колонка 10 — грубо ×2 к области (рейд), не отдельная ротация.",
    ])

    d1 = rank_desc(dps, lambda x: x[1][1]["dmg_st"])
    d5 = rank_desc(dps, lambda x: x[1][5]["dmg_aoe"])
    d10 = rank_desc(dps, lambda x: x[1][10]["dmg_aoe"])
    rows.append([
        "ДД", "Кто лучше бьёт",
        names([(k, r[1]["dmg_st"], 0, 0) for k, r in d1]),
        names([(k, r[5]["dmg_aoe"], 0, 0) for k, r in d5]),
        names([(k, r[10]["dmg_aoe"], 0, 0) for k, r in d10]),
        "1 = фокус. 5 и 10 = сумма по всем целям.",
    ])

    write_sheet(ws, rows[0], rows[1:], [12, 28, 42, 42, 42, 48], [TANK_C, TANK_C, TANK_C, TANK_C, HEAL_C, HEAL_C, DPS_C])
    for i in range(2, 9):
        ws.row_dimensions[i].height = 72

    # ── Танки ──
    ws = wb.create_sheet("Танки")
    headers = [
        "Класс", "Спек", "HP", "Броня+сейв %",
        "Вход 1", "Самохил 1", "Соло 1", "С хилом 1",
        "Вход 5", "Самохил 5", "Соло 5", "С хилом 5",
        "Вход 10", "Самохил 10", "Соло 10", "С хилом 10",
        "Слабое место",
    ]
    trows = []
    for k, r in tanks:
        rec = [k["className"], k["specName"], k["stats"]["hp"], round(r[1]["dr"] * 100, 1)]
        for n in (1, 5, 10):
            inc = tank_incoming(r[n]["dr"], n)
            sh = r[n]["self_heal"]
            solo = (k["stats"]["hp"] + sh) / max(1.0, inc)
            withh = (k["stats"]["hp"] + sh + avg_healer_st) / max(1.0, inc)
            rec += [round(inc, 1), round(sh, 1), round(solo, 3), round(withh, 3)]
        rec.append(weakness_text(k, {**r, 1: {**r[1], "incoming": tank_incoming(r[1]["dr"], 1)},
                                       5: {**r[5], "incoming": tank_incoming(r[5]["dr"], 5)},
                                       10: {**r[10], "incoming": tank_incoming(r[10]["dr"], 10)}}))
        trows.append(rec)
    trows.sort(key=lambda x: -x[6])
    write_sheet(ws, headers, trows, [14, 16, 8, 12, 10, 10, 9, 10, 10, 10, 9, 10, 11, 11, 9, 10, 55], [TANK_C] * len(trows))
    if trows:
        ws.conditional_formatting.add("G2:G20", ColorScaleRule(start_type="min", start_color="F4D6D0", mid_type="percentile", mid_value=50, mid_color="FFF3B0", end_type="max", end_color="C5E0B4"))
        ws.conditional_formatting.add("K2:K20", ColorScaleRule(start_type="min", start_color="F4D6D0", mid_type="percentile", mid_value=50, mid_color="FFF3B0", end_type="max", end_color="C5E0B4"))
        ws.conditional_formatting.add("O2:O20", ColorScaleRule(start_type="min", start_color="F4D6D0", mid_type="percentile", mid_value=50, mid_color="FFF3B0", end_type="max", end_color="C5E0B4"))

    # ── Честный сейв (ниже старой таблицы) ──
    honest_headers = [
        "Класс", "Спек", "HP",
        "Вход сырой", "Вход в HP", "Срезали %",
        "Щит съел", "Самохил", "Шат тик", "Шат остаток",
        "Конец HP", "Умер", "Соло", "С хилом", "Конец HP с хилом", "Умер с хилом",
        "Ротация соло",
    ]
    note_row = 8
    ws.cell(note_row, 1, "Ниже — честный сейв. Танк жмёт стены, блок, самохил, броню. Вход в HP = что сняли с полоски. Старая таблица сверху не менялась.")
    ws.cell(note_row, 1).font = FONT_B
    ws.merge_cells(start_row=note_row, start_column=1, end_row=note_row, end_column=8)

    # компактный индекс — те же Соло / С хилом, что сверху, но из честного прогона
    idx_headers = [
        "Класс", "Спек",
        "Соло 1", "С хилом 1",
        "Соло 5", "С хилом 5",
        "Соло 10", "С хилом 10",
        "Срезали 1 %", "Срезали 5 %", "Срезали 10 %",
        "Умер 1", "Умер 5", "Умер 10",
    ]
    start = 10
    ws.cell(start, 1, "Честный сейв — индекс")
    ws.cell(start, 1).font = FONT_T
    ws.merge_cells(start_row=start, start_column=1, end_row=start, end_column=6)
    start += 1
    for ci, h in enumerate(idx_headers, 1):
        cell = ws.cell(start, ci, h)
        cell.fill = fill("1C3D2A")
        cell.font = FONT_H
        cell.alignment = Alignment(wrap_text=True, vertical="center")
        cell.border = THIN
    ws.row_dimensions[start].height = 28
    start += 1
    idx_rows = []
    for k, _r in tanks:
        rec = [k["className"], k["specName"]]
        for n in (1, 5, 10):
            rec.append(resist[k["key"]][n]["solo"])
            rec.append(resist_h[k["key"]][n]["solo"])
        for n in (1, 5, 10):
            rec.append(resist[k["key"]][n]["cut_pct"])
        for n in (1, 5, 10):
            rec.append(resist[k["key"]][n]["died"])
        idx_rows.append(rec)
    idx_rows.sort(key=lambda x: -x[2])
    idx_first = start
    for rec in idx_rows:
        for ci, v in enumerate(rec, 1):
            cell = ws.cell(start, ci, v)
            cell.font = FONT
            cell.alignment = WRAP
            cell.border = THIN
            cell.fill = fill("E3F0E8")
        start += 1
    idx_last = start - 1
    if idx_last >= idx_first:
        for col in ("C", "D", "E", "F", "G", "H"):
            ws.conditional_formatting.add(
                f"{col}{idx_first}:{col}{idx_last}",
                ColorScaleRule(
                    start_type="min", start_color="F4D6D0",
                    mid_type="percentile", mid_value=50, mid_color="FFF3B0",
                    end_type="max", end_color="C5E0B4",
                ),
            )
    start += 1
    ws.cell(start, 1, "Порядок соло: " + honest_names(1, resist) + "  ·  5: " + honest_names(5, resist) + "  ·  10: " + honest_names(10, resist))
    ws.cell(start, 1).font = FONT_B
    ws.merge_cells(start_row=start, start_column=1, end_row=start, end_column=8)
    start += 2
    for n, title in ((1, "Честный сейв — 1 цель"), (5, "Честный сейв — 5 целей"), (10, "Честный сейв — 10 целей")):
        ws.cell(start, 1, title)
        ws.cell(start, 1).font = FONT_T
        ws.merge_cells(start_row=start, start_column=1, end_row=start, end_column=6)
        start += 1
        for ci, h in enumerate(honest_headers, 1):
            cell = ws.cell(start, ci, h)
            cell.fill = fill("1C3D2A")
            cell.font = FONT_H
            cell.alignment = Alignment(wrap_text=True, vertical="center")
            cell.border = THIN
        ws.row_dimensions[start].height = 28
        start += 1
        block = []
        for k, _r in tanks:
            s = resist[k["key"]][n]
            h = resist_h[k["key"]][n]
            block.append([
                k["className"], k["specName"], k["stats"]["hp"],
                s["raw"], s["hp_in"], s["cut_pct"],
                s["shield_ate"], s["self_heal"], s["stagger_tick"], s["stagger_left"],
                s["end_hp"], s["died"],
                s["solo"], h["solo"],
                h["end_hp"], h["died"],
                s["log"],
            ])
        block.sort(key=lambda x: -x[12])
        for rec in block:
            for ci, v in enumerate(rec, 1):
                cell = ws.cell(start, ci, v)
                cell.font = FONT
                cell.alignment = WRAP
                cell.border = THIN
                cell.fill = fill("E3F0E8")
            start += 1
        start += 2

    for i, w in enumerate([14, 16, 8, 12, 12, 11, 11, 11, 10, 12, 11, 10, 9, 10, 16, 14, 72], 1):
        ws.column_dimensions[get_column_letter(i)].width = max(
            ws.column_dimensions[get_column_letter(i)].width or 0, w
        )

    # ── Хилы ──
    ws = wb.create_sheet("Хилы")
    headers = ["Класс", "Спек", "СТ 12х", "АоЕ-5 12х", "АоЕ-10 грубо", "СТ за ход", "АоЕ-5 за ход", "Ротация (1 цель)", "Слабое место"]
    hrows = []
    for k, r in healers:
        hrows.append([
            k["className"], k["specName"],
            r[1]["heal_st"], r[5]["heal_aoe"], round(r[10]["heal_aoe"] * 2, 1),
            round(r[1]["heal_st"] / HORIZON, 1), round(r[5]["heal_aoe"] / HORIZON, 1),
            r[1]["log"],
            weakness_text(k, r),
        ])
    hrows.sort(key=lambda x: -x[2])
    write_sheet(ws, headers, hrows, [14, 16, 12, 12, 14, 12, 14, 70, 50], [HEAL_C] * len(hrows))

    honest_h = {}
    for k, _r in healers:
        honest_h[k["key"]] = {
            1: simulate_healer_honest(k, 1),
            5: simulate_healer_honest(k, 5),
        }

    note_row = 9
    ws.cell(note_row, 1, "Ниже — честный хил: искусность, крит, хоты, Искупление 55%, Выбор света, змея и эхо Ткача, цепь −5%, Энергия Света. Старая таблица сверху не менялась.")
    ws.cell(note_row, 1).font = FONT_B
    ws.merge_cells(start_row=note_row, start_column=1, end_row=note_row, end_column=8)

    idx_h = [
        "Класс", "Спек",
        "СТ 12х", "СТ за ход",
        "АоЕ-5 12х", "АоЕ-5 за ход",
        "АоЕ-10 грубо", "АоЕ / СТ",
        "Ротация СТ", "Ротация область",
    ]
    start = 11
    ws.cell(start, 1, "Честный хил — индекс")
    ws.cell(start, 1).font = FONT_T
    ws.merge_cells(start_row=start, start_column=1, end_row=start, end_column=6)
    start += 1
    for ci, h in enumerate(idx_h, 1):
        cell = ws.cell(start, ci, h)
        cell.fill = fill("1C3D2A")
        cell.font = FONT_H
        cell.alignment = Alignment(wrap_text=True, vertical="center")
        cell.border = THIN
    ws.row_dimensions[start].height = 28
    start += 1
    idx_first = start
    block = []
    for k, _r in healers:
        s = honest_h[k["key"]][1]
        a = honest_h[k["key"]][5]
        block.append([
            k["className"], k["specName"],
            s["heal_st"], round(s["heal_st"] / HORIZON, 1),
            a["heal_aoe"], round(a["heal_aoe"] / HORIZON, 1),
            round(a["heal_aoe"] * 2, 1),
            round(a["heal_aoe"] / max(1.0, s["heal_st"]), 2),
            s["log"], a["log"],
        ])
    block.sort(key=lambda x: -x[2])
    for rec in block:
        for ci, v in enumerate(rec, 1):
            cell = ws.cell(start, ci, v)
            cell.font = FONT
            cell.alignment = WRAP
            cell.border = THIN
            cell.fill = fill("E3F0E8")
        start += 1
    idx_last = start - 1
    if idx_last >= idx_first:
        for col in ("C", "D", "E", "F", "G"):
            ws.conditional_formatting.add(
                f"{col}{idx_first}:{col}{idx_last}",
                ColorScaleRule(
                    start_type="min", start_color="F4D6D0",
                    mid_type="percentile", mid_value=50, mid_color="FFF3B0",
                    end_type="max", end_color="C5E0B4",
                ),
            )
    start += 1
    st_rank = sorted(healers, key=lambda x: -honest_h[x[0]["key"]][1]["heal_st"])
    ao_rank = sorted(healers, key=lambda x: -honest_h[x[0]["key"]][5]["heal_aoe"])
    ws.cell(start, 1, "Порядок СТ: " + " → ".join(
        f"{i+1}. {k['className']} {k['specName']}" for i, (k, _) in enumerate(st_rank)
    ) + "  ·  область: " + " → ".join(
        f"{i+1}. {k['className']} {k['specName']}" for i, (k, _) in enumerate(ao_rank)
    ))
    ws.cell(start, 1).font = FONT_B
    ws.merge_cells(start_row=start, start_column=1, end_row=start, end_column=8)
    for i, w in enumerate([14, 16, 12, 12, 12, 14, 14, 10, 62, 62], 1):
        ws.column_dimensions[get_column_letter(i)].width = max(
            ws.column_dimensions[get_column_letter(i)].width or 0, w
        )

    # ── ДД ──
    ws = wb.create_sheet("ДД")
    headers = [
        "Класс", "Спек", "СТ 1", "Сумма 5", "Сумма 10",
        "СТ за ход", "5 / СТ", "10 / 5",
        "Ротация (1)", "Ротация (5)", "Слабое место",
    ]
    drows = []
    for k, r in dps:
        st = r[1]["dmg_st"]
        a5 = r[5]["dmg_aoe"]
        a10 = r[10]["dmg_aoe"]
        drows.append([
            k["className"], k["specName"],
            st, a5, a10,
            round(st / HORIZON, 1),
            round(a5 / max(1, st), 2),
            round(a10 / max(1, a5), 2),
            r[1]["log"], r[5]["log"],
            weakness_text(k, r),
        ])
    drows.sort(key=lambda x: -x[2])
    write_sheet(ws, headers, drows, [14, 18, 10, 10, 10, 10, 9, 9, 52, 52, 48], [DPS_C] * len(drows))
    if drows:
        ws.conditional_formatting.add("C2:E40", ColorScaleRule(start_type="min", start_color="F4D6D0", mid_type="percentile", mid_value=50, mid_color="FFF3B0", end_type="max", end_color="C5E0B4"))

    # ── Слабые стороны ──
    ws = wb.create_sheet("Слабые стороны")
    headers = ["Роль", "Класс", "Спек", "Главная дыра", "Что смотреть в бою"]
    wrows = []
    tips = {
        "tank": "В 5+ смотри входящий, не урон. Стены не в индексе пиком.",
        "healer": "СТ vs область часто разные кнопки. Мана на длинном пулле.",
        "dps": "Если 5/СТ < 1.5 — это чисто фокус. Если 10/5 ≈ 1 — область уже выжата.",
    }
    for k, r in pack:
        wrows.append([k["role"], k["className"], k["specName"], weakness_text(k, r), tips.get(k["role"], "")])
    write_sheet(ws, headers, wrows, [10, 16, 18, 70, 40])

    # ── Ротации ──
    ws = wb.create_sheet("Ротации")
    headers = ["Класс", "Спек", "1 цель", "5 целей", "10 целей"]
    rrows = [[k["className"], k["specName"], r[1]["log"], r[5]["log"], r[10]["log"]] for k, r in pack]
    write_sheet(ws, headers, rrows, [14, 18, 70, 70, 70])

    wb.save(OUT)
    print("wrote", OUT)


if __name__ == "__main__":
    main()
