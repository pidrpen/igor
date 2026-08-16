# -*- coding: utf-8 -*-
"""Дописать в существующий обзор учёт ресурса. Жёлтое «Задача мне» не трогает."""
from __future__ import annotations

import re
from collections import defaultdict
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parents[1]
XLSX = ROOT / "баланс_полный_обзор.xlsx"

INK = "1A1A1A"
HEAD = "2C2416"
HEAD_F = "F7F1E3"
FACT = "E8F0E3"
PROP = "E3EEF7"
YEL = "FFF3B0"
OK = "E8F0E3"
WARN = "F8E6C8"
BAD = "F4D6D0"
THIN = Border(
    left=Side(style="thin", color="D0C4A8"),
    right=Side(style="thin", color="D0C4A8"),
    top=Side(style="thin", color="D0C4A8"),
    bottom=Side(style="thin", color="D0C4A8"),
)
WRAP = Alignment(wrap_text=True, vertical="top")
FONT = Font(name="Calibri", size=10, color=INK)
FONT_H = Font(name="Calibri", size=10, bold=True, color=HEAD_F)
FONT_B = Font(name="Calibri", size=10, bold=True, color=INK)


def fill(h):
    return PatternFill("solid", fgColor=h)


def parse_cost(s):
    s = (s or "").strip()
    c = g = cs = gs = rp = 0
    runes = 0
    if not s or s == "0":
        return dict(c=0, g=0, cs=0, gs=0, rp=0, runes=0)
    m = re.search(r"c:(\d+(?:\.\d+)?)", s)
    if m:
        c = float(m.group(1))
    m = re.search(r"(?<![a-z])g:(\d+(?:\.\d+)?)", s)
    if m:
        g = float(m.group(1))
    m = re.search(r"cs:(\d+(?:\.\d+)?)", s)
    if m:
        cs = float(m.group(1))
    m = re.search(r"gs:(\d+(?:\.\d+)?)", s)
    if m:
        gs = float(m.group(1))
    m = re.search(r"\+(\d+)\s*RP", s, re.I)
    if m:
        rp = float(m.group(1))
    if "рун" in s.lower():
        runes = 1
        for m in re.finditer(r"×(\d+)", s):
            runes = max(runes, int(m.group(1)))
    return dict(c=c, g=g, cs=cs, gs=gs, rp=rp, runes=runes)


def role_of(p):
    spend = (p["c"] > 0) or (p["cs"] > 0)
    build = (p["g"] > 0) or (p["gs"] > 0) or (p["rp"] > 0) or (p["runes"] > 0 and p["cs"] == 0)
    if spend and build:
        return "смесь"
    if spend:
        return "тратит"
    if build:
        return "копит"
    return "бесплатно"


def style_head(cell):
    cell.fill = fill(HEAD)
    cell.font = FONT_H
    cell.alignment = WRAP
    cell.border = THIN


def style_cell(cell, bg=FACT):
    cell.fill = fill(bg)
    cell.font = FONT
    cell.alignment = WRAP
    cell.border = THIN


def main():
    wb = load_workbook(XLSX)
    kn = wb["Кнопки"]
    # existing: 1-21, 21 = Задача мне — не трогать
    heads = [
        (22, "копит/тратит"),
        (23, "цена осн."),
        (24, "ген осн."),
        (25, "цена втор."),
        (26, "ген втор./RP"),
        (27, "т за 10 цены"),
        (28, "флаг экономики"),
    ]
    for col, title in heads:
        cell = kn.cell(4, col, title)
        style_head(cell)
        kn.column_dimensions[get_column_letter(col)].width = 16 if col < 28 else 36

    # group damage-ish rows by spec
    spec_rows = defaultdict(list)
    for r in range(5, kn.max_row + 1):
        cls = kn.cell(r, 1).value
        spec = kn.cell(r, 2).value
        typ = (kn.cell(r, 7).value or "")
        mid = kn.cell(r, 19).value
        cost_s = kn.cell(r, 8).value
        parsed = parse_cost(cost_s)
        eco = role_of(parsed)
        kn.cell(r, 22, eco)
        kn.cell(r, 23, parsed["c"] or None)
        kn.cell(r, 24, parsed["g"] or None)
        kn.cell(r, 25, parsed["cs"] or None)
        kn.cell(r, 26, (parsed["gs"] or parsed["rp"] or None))
        mid_n = float(mid) if isinstance(mid, (int, float)) else None
        per10 = None
        if mid_n is not None and parsed["c"] > 0:
            per10 = round(mid_n * 10 / parsed["c"], 1)
        elif mid_n is not None and parsed["cs"] > 0:
            per10 = round(mid_n * 10 / parsed["cs"], 1)
        kn.cell(r, 27, per10)
        bg = FACT
        if eco == "тратит":
            bg = PROP
        elif eco == "копит":
            bg = OK
        elif eco == "бесплатно":
            bg = WARN
        for c in range(22, 29):
            style_cell(kn.cell(r, c), bg)
        if typ in ("damage", "aoe", "dot") and mid_n is not None:
            spec_rows[(cls, spec)].append({
                "row": r, "name": kn.cell(r, 6).value, "eco": eco,
                "mid": mid_n, "c": parsed["c"], "cs": parsed["cs"],
                "id": kn.cell(r, 5).value,
            })

    spec_flags = {}
    for key, rows in spec_rows.items():
        builders = [x for x in rows if x["eco"] in ("копит", "смесь") and x["c"] == 0]
        spenders = [x for x in rows if x["eco"] == "тратит" and (x["c"] > 0 or x["cs"] > 0)]
        free_fat = [x for x in rows if x["eco"] == "бесплатно" and x["mid"] >= 20]
        notes = []
        if builders and spenders:
            bmax = max(builders, key=lambda x: x["mid"])
            smax = max(spenders, key=lambda x: x["mid"])
            if bmax["mid"] >= smax["mid"] - 0.5:
                notes.append(
                    f"копилка «{bmax['name']}» {bmax['mid']}т ≥ траты «{smax['name']}» {smax['mid']}т — смотреть"
                )
            else:
                notes.append(
                    f"трата «{smax['name']}» {smax['mid']}т > копилки «{bmax['name']}» {bmax['mid']}т — норма"
                )
        if free_fat:
            notes.append("бесплатно и жирно: " + ", ".join(f"{x['name']} {x['mid']}т" for x in free_fat))
        flag = "; ".join(notes) if notes else ""
        spec_flags[key] = flag
        for x in rows:
            kn.cell(x["row"], 28, flag)
            style_cell(kn.cell(x["row"], 28), WARN if "смотреть" in flag or "бесплатно" in flag else OK)

    # Спеки: колонка N только если Задача (M) пустая
    sp = wb["Спеки"]
    nhead = sp.cell(4, 14, "учёт ресурса")
    style_head(nhead)
    sp.column_dimensions["N"].width = 42
    # map russian spec name -> flag via Кнопки class+spec
    for r in range(5, 42):
        sid = sp.cell(r, 1).value
        cls = sp.cell(r, 2).value
        spec = sp.cell(r, 3).value
        task = sp.cell(r, 13).value
        flag = spec_flags.get((cls, spec), "")
        cell = sp.cell(r, 14)
        if task and str(task).strip():
            # уже комментировали — только справка, не переписывать задачу
            if not cell.value:
                cell.value = "(уже есть задача; ресурс: " + (flag or "см. Кнопки") + ")"
            style_cell(cell, PROP)
        else:
            cell.value = flag or "нет урона для сравнения копилка/трата"
            style_cell(cell, WARN if flag and ("смотреть" in flag or "бесплатно" in flag) else OK)

    # маленький лист-сводка только незакомментированных
    if "Ресурс" in wb.sheetnames:
        del wb["Ресурс"]
    sm = wb.create_sheet("Ресурс", 4)
    sm["A1"] = "Оценка с учётом трат (не пересчитывает голые т). Жёлтые задачи не тронуты."
    sm["A2"] = (
        "Правило: кто тратит ресурс, может бить жирнее копилки. "
        "Голые т на листе Кнопки ресурс не считают. ДК: руна копит силу рун, Лик/Удар льда тратят её."
    )
    sm.merge_cells("A1:F1")
    sm.merge_cells("A2:F2")
    headers = ["id", "Класс", "Спек", "задача?", "флаг экономики", "что смотреть"]
    for i, h in enumerate(headers, 1):
        style_head(sm.cell(4, i, h))
    rr = 5
    for r in range(5, 42):
        sid = wb["Спеки"].cell(r, 1).value
        cls = wb["Спеки"].cell(r, 2).value
        spec = wb["Спеки"].cell(r, 3).value
        task = wb["Спеки"].cell(r, 13).value
        commented = bool(task and str(task).strip())
        flag = spec_flags.get((cls, spec), "")
        if commented:
            continue
        sm.cell(rr, 1, sid)
        sm.cell(rr, 2, cls)
        sm.cell(rr, 3, spec)
        sm.cell(rr, 4, "нет")
        sm.cell(rr, 5, flag)
        look = ""
        if flag and "смотреть" in flag:
            look = "копилка не слабее траты"
        elif flag and "бесплатно" in flag:
            look = "есть жирный бесплатный удар"
        else:
            look = "трата жирнее копилки — не резать из-за голых т"
        sm.cell(rr, 6, look)
        bg = WARN if look != "трата жирнее копилки — не резать из-за голых т" else OK
        for c in range(1, 7):
            style_cell(sm.cell(rr, c), bg)
        rr += 1
    for col, w in enumerate([8, 16, 18, 10, 55, 36], 1):
        sm.column_dimensions[get_column_letter(col)].width = w
    sm.row_dimensions[1].height = 22
    sm.row_dimensions[2].height = 36
    sm["A1"].font = FONT_B
    sm["A2"].alignment = WRAP

    wb.save(XLSX)
    print("patched", XLSX)
    print("uncommented specs on Ресурс:", rr - 5)


if __name__ == "__main__":
    main()
